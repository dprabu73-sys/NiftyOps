"""
End-to-end self-test:
1. Check session + JWT status
2. Trigger headless login if needed
3. Run extraction for last 2 trading days
4. Poll until complete
5. Verify Excel output
"""
import requests
import json
import time
import os

BASE = "http://127.0.0.1:5005"

def step(msg):
    print(f"\n{'='*60}")
    print(f"  {msg}")
    print('='*60)

# ── Step 1: Check session ──────────────────────────────────────
step("STEP 1: Checking session status")
r = requests.get(f"{BASE}/api/session-status", timeout=10)
status = r.json()
print(f"Session: {status}")
has_session = status.get("has_session", False)

# Check if JWT exists in .env
env_path = r"c:\Users\dprab\OneDrive\Desktop\Tradingview\Prabu New Trading view\.env"
has_jwt = False
tv_session = None
with open(env_path) as f:
    for line in f:
        if line.startswith("TV_JWT_TOKEN="):
            val = line.split("=", 1)[1].strip()
            if val and val.startswith("eyJ"):
                has_jwt = True
        if line.startswith("TV_SESSION_ID="):
            tv_session = line.split("=", 1)[1].strip()

print(f"Has session cookie: {has_session}")
print(f"Has JWT token: {has_jwt}")

# ── Step 2: Auto-login if no JWT ───────────────────────────────
if not has_jwt:
    step("STEP 2: Triggering auto-login to capture JWT")
    r2 = requests.post(f"{BASE}/api/headless-login",
                       json={"username": "brokerworkflowhub", "password": "Nithik@20252"},
                       timeout=10)
    print(f"Login trigger: {r2.json()}")

    # Poll login status for up to 2 minutes
    deadline = time.time() + 120
    while time.time() < deadline:
        ls = requests.get(f"{BASE}/api/headless-login-status", timeout=5).json()
        print(f"  Login: [{ls.get('status')}] {ls.get('message','')[:80]}")
        if ls.get("status") in ("success", "failed"):
            break
        time.sleep(3)

    # Re-check JWT
    with open(env_path) as f:
        for line in f:
            if line.startswith("TV_JWT_TOKEN="):
                val = line.split("=", 1)[1].strip()
                if val and val.startswith("eyJ"):
                    has_jwt = True
    print(f"JWT after login: {has_jwt}")
else:
    step("STEP 2: JWT already present — skipping auto-login")
    print("JWT token found in .env. Proceeding to extraction.")

# ── Step 3: Run extraction for last 2 trading days ─────────────
step("STEP 3: Starting extraction (last 2 trading days)")
extract_payload = {
    "symbol": "NIFTY",
    "exchange": "NSE",
    "interval": "5 Minutes",
    "n_bars": 800,
    "filename": "SelfTest_2Days",
    "time_filter": "last5",
    "strike_offset": 100,
    "baseline_interval": "15 Minutes",
    "signal_interval": "5 Minutes",
    "session_id": tv_session or ""
}
r3 = requests.post(f"{BASE}/api/extract", json=extract_payload, timeout=15)
resp3 = r3.json()
print(f"Extract response: {resp3}")
task_id = resp3.get("task_id")
if not task_id:
    print("ERROR: No task_id returned. Aborting.")
    exit(1)

# ── Step 4: Poll task until complete ──────────────────────────
step(f"STEP 4: Polling task {task_id[:8]}...")
deadline = time.time() + 600  # 10 min max
last_log_count = 0
while time.time() < deadline:
    tr = requests.get(f"{BASE}/api/task/{task_id}", timeout=10).json()
    logs = tr.get("logs", [])
    # Print new logs
    for log in logs[last_log_count:]:
        icon = {"info": "ℹ", "success": "✅", "warn": "⚠️", "error": "❌"}.get(log.get("type",""), "•")
        print(f"  {icon} {log.get('message','')[:100]}")
    last_log_count = len(logs)

    task_status = tr.get("status", "")
    if task_status in ("completed", "failed"):
        print(f"\n  Final status: {task_status.upper()}")
        break
    time.sleep(3)

# ── Step 5: Verify Excel output ───────────────────────────────
step("STEP 5: Verifying Excel output")
if task_status == "completed":
    # Download the Excel
    dl = requests.get(f"{BASE}/api/download/{task_id}", timeout=30)
    if dl.status_code == 200:
        out_path = r"c:\Users\dprab\OneDrive\Desktop\Tradingview\Prabu New Trading view\temp_exports\SelfTest_Output.xlsx"
        with open(out_path, "wb") as f:
            f.write(dl.content)
        size_kb = len(dl.content) / 1024
        print(f"  ✅ Excel downloaded: {size_kb:.1f} KB → {out_path}")

        # Verify with openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(out_path)
            print(f"  ✅ Workbook sheets: {wb.sheetnames}")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.max_row
                cols = ws.max_column
                # Count non-empty cells
                filled = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
                print(f"     [{sheet_name}]: {rows} rows × {cols} cols, {filled} filled cells")
        except Exception as e:
            print(f"  ⚠️ Could not verify with openpyxl: {e}")
    else:
        print(f"  ❌ Download failed: HTTP {dl.status_code}")
else:
    print(f"  ❌ Task did not complete successfully (status={task_status})")
    # Print last 10 logs
    tr = requests.get(f"{BASE}/api/task/{task_id}", timeout=10).json()
    for log in tr.get("logs", [])[-10:]:
        print(f"    {log.get('type','').upper()}: {log.get('message','')}")

print("\n" + "="*60)
print("  SELF-TEST COMPLETE")
print("="*60)
