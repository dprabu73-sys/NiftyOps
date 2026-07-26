import os
import sys
import uuid
import time
import logging
import threading
import webbrowser
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from flask import Flask, request, jsonify, render_template, send_file
from tvDatafeed import TvDatafeed, Interval
if getattr(sys, 'frozen', False):
    # Running inside PyInstaller bundle (EXE)
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
else:
    # Running in normal python environment
    template_folder = 'templates'
    static_folder = 'static'

app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
# In-memory storage for active tasks
tasks = {}
tasks_lock = threading.Lock()

# Directory for caching Excel downloads
CACHE_DIR = os.path.join(os.getcwd(), 'temp_exports')
os.makedirs(CACHE_DIR, exist_ok=True)

# Map UI interval labels to tvDatafeed Interval Enums
INTERVAL_MAP = {
    "1 Minute": Interval.in_1_minute,
    "3 Minutes": Interval.in_3_minute,
    "5 Minutes": Interval.in_5_minute,
    "15 Minutes": Interval.in_15_minute,
    "30 Minutes": Interval.in_30_minute,
    "45 Minutes": Interval.in_45_minute,
    "1 Hour": Interval.in_1_hour,
    "2 Hours": Interval.in_2_hour,
    "3 Hours": Interval.in_3_hour,
    "4 Hours": Interval.in_4_hour,
    "Daily": Interval.in_daily,
    "Weekly": Interval.in_weekly,
    "Monthly": Interval.in_monthly
}

def log_task(task_id, message, msg_type="info"):
    with tasks_lock:
        if task_id in tasks:
            tasks[task_id]['logs'].append({
                'time': datetime.now().strftime('%H:%M:%S'),
                'message': message,
                'type': msg_type
            })


# Ã¢â€â‚¬Ã¢â€â‚¬ TvDatafeed connection pool Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
# Each parallel worker thread keeps its own tv instance so WebSocket connections
# are NOT recreated on every fetch (the main cause of the 5-28s delay).
_tv_pool_lock = threading.Lock()
_tv_pool: dict = {}          # maps thread-id Ã¢â€ â€™ TvDatafeed instance

def _fetch_jwt_from_session(session_id: str) -> str:
    """Try to fetch a real JWT auth token from TradingView using the session cookie.
    TradingView embeds the JWT in the HTML of the main page for logged-in users.
    Returns the JWT string, or empty string if not found.
    """
    import re as _re
    import requests as _requests
    try:
        cookie_str = f"sessionid={session_id}"
        headers = {
            "Cookie": cookie_str,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.tradingview.com/",
            "Accept": "text/html,application/xhtml+xml",
        }
        resp = _requests.get("https://www.tradingview.com/", headers=headers, timeout=15)
        if resp.status_code != 200:
            return ""
        # JWT is embedded as: "auth_token":"eyJ..."
        match = _re.search(r'"auth_token"\s*:\s*"([A-Za-z0-9_\-\.]+)"', resp.text)
        if match:
            jwt = match.group(1)
            logging.info(f"JWT auto-fetched from TradingView session (len={len(jwt)})")
            return jwt
    except Exception as e:
        logging.warning(f"JWT auto-fetch failed: {e}")
    return ""


def _get_pooled_tv(username=None, password=None, tv_session=None):
    """Return the TvDatafeed instance for the current thread, creating it once.

    Auth flow (in priority order):
      1. TV_JWT_TOKEN in .env  — direct JWT, best quality, works for NFO options
      2. Session cookie + auto-fetched JWT from TradingView HTML
      3. Session cookie alone with unauthorized_user_token (index data only)
      4. Username/password login fallback

    NFO options REQUIRE a valid JWT in set_auth_token — the session cookie alone
    is not sufficient. Use the Settings page to paste your TV_JWT_TOKEN.
    """
    import json as _json
    tid = threading.get_ident()
    with _tv_pool_lock:
        if tid not in _tv_pool:
            # --- Unmask preview tokens ----------------------------------------
            if tv_session and ('***' in tv_session or tv_session == '***saved***'):
                tv_session = ""

            # --- Load credentials from .env -----------------------------------
            tv_jwt = ""
            env_path = os.path.join(os.path.dirname(__file__), '.env')
            if os.path.exists(env_path):
                try:
                    with open(env_path, 'r', encoding='utf-8') as ef:
                        for eline in ef:
                            eline = eline.strip()
                            if not tv_session and eline.startswith('TV_SESSION_ID='):
                                tv_session = eline.split('=', 1)[1].strip()
                            elif eline.startswith('TV_JWT_TOKEN='):
                                tv_jwt = eline.split('=', 1)[1].strip()
                except Exception:
                    pass

            if tv_session and tv_session.strip():
                raw_session = tv_session.strip()
                cookie_val = raw_session
                if cookie_val.startswith('sessionid='):
                    cookie_val = cookie_val[len('sessionid='):]
                cookie_str = f"sessionid={cookie_val}"

                # If no JWT in .env, try to auto-fetch from TradingView page
                if not tv_jwt:
                    tv_jwt = _fetch_jwt_from_session(cookie_val)
                    if tv_jwt:
                        # Cache it to .env so we don't refetch every time
                        try:
                            with open(env_path, 'a', encoding='utf-8') as ef:
                                ef.write(f"\nTV_JWT_TOKEN={tv_jwt}\n")
                            logging.info("TV_JWT_TOKEN auto-saved to .env")
                        except Exception:
                            pass

                # Create a no-login instance (guest skeleton)
                tv_new = TvDatafeed()

                # Patch layer-1: inject Cookie into WS handshake headers
                ws_headers = _json.dumps({
                    "Origin": "https://data.tradingview.com",
                    "Cookie": cookie_str
                })
                tv_new._TvDatafeed__ws_headers = ws_headers

                # Patch layer-2: set_auth_token
                # Use real JWT if available (needed for NFO options)
                # Fall back to unauthorized_user_token only for index data
                tv_new.token = tv_jwt if tv_jwt else 'unauthorized_user_token'
                if tv_jwt:
                    logging.info(f"TV auth: using real JWT token (len={len(tv_jwt)}) + session cookie")
                else:
                    logging.warning("TV auth: using unauthorized_user_token — NFO options may be unavailable. "
                                    "Add TV_JWT_TOKEN to .env for full derivatives access.")
            else:
                tv_new = TvDatafeed(username, password)

            _tv_pool[tid] = tv_new
        return _tv_pool[tid]


def _flush_tv_pool():
    """Clear all pooled TvDatafeed instances so the next fetch creates fresh ones.
    Call this whenever a new session token is saved to .env."""
    with _tv_pool_lock:
        _tv_pool.clear()


def safe_get_hist(tv, symbol, exchange, interval, n_bars, retries=2, delay=1.0):
    """Fetch historical bars with minimal retries for speed."""
    for attempt in range(retries):
        try:
            data = tv.get_hist(symbol=symbol, exchange=exchange, interval=interval, n_bars=n_bars)
            if data is not None and not data.empty:
                return data
        except Exception:
            pass
        if attempt < retries - 1:
            time.sleep(delay)
    return None



def extraction_thread(task_id, username, password, symbol, exchange, interval_name, n_bars, filename, time_filter='all', strike_offset=100, google_sheet_url=None, live_today_only=False, tv_session=None, baseline_interval_name='15 Minutes', signal_interval_name='5 Minutes'):
    def _run_simulation_logic(candles_df, prev_day_1515_close_ha, today_0915_close_ha, target_points=25, sl_type='close'):
        if candles_df is None or candles_df.empty:
            return {
                'baseline': '', 'entry_price': '', 'entry_time': '',
                'target_val': '', 'sl_val': '', 'exit_reason': 'No Data',
                'exit_time': '', 'exit_price': '', 'pnl': ''
            }
        try:
            prev_ha = float(prev_day_1515_close_ha) if prev_day_1515_close_ha != "" else 0.0
            today_ha = float(today_0915_close_ha) if today_0915_close_ha != "" else 0.0
        except Exception:
            prev_ha = 0.0
            today_ha = 0.0
            
        computed_baseline = max(prev_ha, today_ha)
        final_baseline = computed_baseline
        
        sorted_df = candles_df.sort_values(by='ist_time')
        candles = []
        for _, row in sorted_df.iterrows():
            try:
                time_str = str(row['ist_time']).strip()
                if ' ' in time_str:
                    time_str = time_str.split(' ')[-1]
                if len(time_str.split(':')) > 2:
                    time_str = ':'.join(time_str.split(':')[:2])
                
                # Slicing Rule: Completely ignore the 9:15 AM and 9:20 AM candles
                if time_str in ['09:15', '09:20']:
                    continue
                    
                candles.append({
                    'time': time_str,
                    'open': float(row['open']),
                    'high': float(row['high']),
                    'low': float(row['low']),
                    'close': float(row['close'])
                })
            except Exception:
                continue
                
        if not candles:
            return {
                'baseline': round(final_baseline, 2), 'entry_price': '', 'entry_time': '',
                'target_val': '', 'sl_val': round(final_baseline, 2), 'exit_reason': 'No Data',
                'exit_time': '', 'exit_price': '', 'pnl': ''
            }
            
        # 1. Setup Candle Scanner: Monitor candles for a candle completely below baseline (High < Baseline)
        setup_candle_idx = None
        locked_sl = None
        for i, c in enumerate(candles):
            if c['high'] < final_baseline:
                setup_candle_idx = i
                locked_sl = c['low']
                break
                
        # 2. Entry Breach Scanner: Scan for 2nd breach of baseline
        first_breach_idx = None
        entry_idx = None
        entry_price = None
        entry_time = None
        
        for i, c in enumerate(candles):
            high_val = max(c['open'], c['high'], c['low'], c['close'])
            if high_val > final_baseline:
                if first_breach_idx is None:
                    first_breach_idx = i
                elif entry_idx is None:
                    entry_idx = i
                    entry_price = c['open']
                    entry_time = c['time']
                    break
                    
        if first_breach_idx is not None and entry_idx is None:
            c = candles[first_breach_idx]
            entry_idx = first_breach_idx
            entry_price = max(c['open'], c['high'], c['low'], c['close'])
            entry_time = c['time']
            
        if entry_idx is None:
            return {
                'baseline': round(final_baseline, 2), 'entry_price': '', 'entry_time': '',
                'target_val': '', 'sl_val': round(locked_sl if locked_sl is not None else final_baseline, 2), 'exit_reason': 'No Entry',
                'exit_time': '', 'exit_price': '', 'pnl': 0.0
            }
            
        # 3. Target Exit Scanner
        entry_ref_high = candles[entry_idx]['high']
        target_val = entry_ref_high + target_points
        target_hit_idx = None
        target_hit_time = None
        target_exit_price = target_val
        
        for i in range(entry_idx, len(candles)):
            c = candles[i]
            if c['high'] >= target_val:
                target_hit_idx = i
                target_hit_time = c['time']
                break
                
        # 4. Stop Loss Exit Scanner: Active ONLY if setup candle has been identified
        sl_ref = locked_sl if locked_sl is not None else final_baseline
        sl_hit_idx = None
        sl_hit_time = None
        sl_exit_price = None
        
        if setup_candle_idx is not None:
            # Check subsequent candles after setup candle has completed
            start_scan_idx = max(entry_idx, setup_candle_idx + 1)
            for i in range(start_scan_idx, len(candles)):
                c = candles[i]
                trigger_val = c['low'] if sl_type == 'low' else c['close']
                if trigger_val < sl_ref:
                    sl_hit_idx = i
                    sl_hit_time = c['time']
                    sl_exit_price = min(c['open'], c['high'], c['low'], c['close'])
                    break
                
        exit_reason = "Open/No Exit"
        exit_time = "Ã¢â‚¬â€"
        exit_price = 0.0
        pnl = 0.0
        
        if target_hit_idx is not None and sl_hit_idx is not None:
            if target_hit_idx < sl_hit_idx:
                exit_reason = "Target Hit"
                exit_time = target_hit_time
                exit_price = target_exit_price
                pnl = exit_price - entry_price
            else:
                exit_reason = "Stop Loss Hit"
                exit_time = sl_hit_time
                exit_price = sl_exit_price
                pnl = exit_price - entry_price
        elif target_hit_idx is not None:
            exit_reason = "Target Hit"
            exit_time = target_hit_time
            exit_price = target_exit_price
            pnl = exit_price - entry_price
        elif sl_hit_idx is not None:
            exit_reason = "Stop Loss Hit"
            exit_time = sl_hit_time
            exit_price = sl_exit_price
            pnl = exit_price - entry_price

        # Calculate dynamic audit values
        first_breach_time = candles[first_breach_idx]['time'] if first_breach_idx is not None else ""
        second_breach_time = candles[entry_idx]['time'] if entry_idx is not None else ""
        entry_candle_high = candles[entry_idx]['high'] if entry_idx is not None else ""
        setup_candle_time = candles[setup_candle_idx]['time'] if setup_candle_idx is not None else ""
        setup_candle_low = candles[setup_candle_idx]['low'] if setup_candle_idx is not None else ""
        sl_activation_time = candles[setup_candle_idx]['time'] if setup_candle_idx is not None else ""

        trade_duration = ""
        if entry_time and exit_time and entry_time not in ("", "—", "Ã¢â‚¬â€ ", "—") and exit_time not in ("", "—", "Ã¢â‚¬â€ ", "—"):
            try:
                from datetime import datetime
                t_entry = datetime.strptime(entry_time, '%H:%M')
                t_exit = datetime.strptime(exit_time, '%H:%M')
                trade_duration = int((t_exit - t_entry).total_seconds() / 60)
            except Exception:
                pass
            
        return {
            'baseline': round(final_baseline, 2),
            'entry_price': round(entry_price, 2) if entry_price else '',
            'entry_time': entry_time or '',
            'target_val': round(target_val, 2) if target_val else '',
            'sl_val': round(sl_ref, 2),
            'exit_reason': exit_reason,
            'exit_time': exit_time,
            'exit_price': round(exit_price, 2) if exit_price else '',
            'pnl': round(pnl, 2),
            'first_breach_time': first_breach_time,
            'second_breach_time': second_breach_time,
            'entry_candle_high': round(entry_candle_high, 2) if entry_candle_high else '',
            'setup_candle_time': setup_candle_time,
            'setup_candle_low': round(setup_candle_low, 2) if setup_candle_low else '',
            'sl_activation_time': sl_activation_time,
            'trade_duration': trade_duration
        }

    try:
        # Check if the submitted session ID is a masked preview token
        if tv_session and ('***' in tv_session or tv_session == '***saved***'):
            tv_session = ""

        # Always read .env to load missing config values
        import os
        env_path = os.path.join(os.path.dirname(__file__), '.env')
        if os.path.exists(env_path):
            try:
                with open(env_path, 'r', encoding='utf-8') as ef:
                    for eline in ef:
                        eline = eline.strip()
                        if eline.startswith('TRADINGVIEW_USERNAME=') or eline.startswith('TV_USERNAME='):
                            if not username:
                                username = eline.split('=', 1)[1].strip()
                        elif eline.startswith('TRADINGVIEW_PASSWORD=') or eline.startswith('TV_PASSWORD='):
                            if not password:
                                password = eline.split('=', 1)[1].strip()
                        elif eline.startswith('TV_SESSION_ID='):
                            if not tv_session:
                                tv_session = eline.split('=', 1)[1].strip()
            except Exception:
                pass

        # Initialize TV client — prefer pre-authenticated session token
        # Check if the submitted session ID is actually a bookmarklet JS code
        is_invalid_token = False
        if tv_session:
            t_str = tv_session.strip()
            if t_str.startswith('javascript:') or 'function(' in t_str or 'document.cookie' in t_str or len(t_str) > 120:
                is_invalid_token = True
                
        if tv_session and tv_session.strip() and not is_invalid_token:
            log_task(task_id, "Using pre-authenticated TradingView session token (bypasses CAPTCHA)...", "info")
            tv = TvDatafeed()
            t_val = tv_session.strip()
            if t_val.startswith("sessionid="):
                t_val = t_val[len("sessionid="):]
            cookie_str = f"sessionid={t_val}"
            
            # Patch layer-1: inject Cookie header
            import json as _json
            ws_headers = _json.dumps({
                "Origin": "https://data.tradingview.com",
                "Cookie": cookie_str
            })
            tv._TvDatafeed__ws_headers = ws_headers
            
            # Patch layer-2: set_auth_token message
            tv.token = 'unauthorized_user_token'
        else:
            if is_invalid_token:
                log_task(task_id, "WARNING: Ignored invalid submitted session token (bookmarklet code detected).", "warn")
            if username and password:
                log_task(task_id, f"Logging in with credentials for user '{username}'...", "info")
                tv = TvDatafeed(username=username, password=password)
                if tv.token == 'unauthorized_user_token':
                    log_task(task_id, "WARNING: Login failed (CAPTCHA block). Add TV_SESSION_ID to .env for full access.", "warn")
            else:
                log_task(task_id, "No credentials specified. Connecting as guest...", "info")
                tv = TvDatafeed()
            
        interval = INTERVAL_MAP.get(interval_name, Interval.in_1_minute)
        
        # Mapped custom intervals for options calculations
        baseline_interval = INTERVAL_MAP.get(baseline_interval_name, Interval.in_15_minute)
        signal_interval = INTERVAL_MAP.get(signal_interval_name, Interval.in_5_minute)
        
        # Enforce 1-Minute interval as HA Extractor requires specific timestamps
        actual_interval = Interval.in_1_minute
        if live_today_only:
            # Enforce at least 1000 bars so we always cover today's open and yesterday's close
            actual_n_bars = max(n_bars, 1000)
            log_task(task_id, f"LIVE MODE: Fetching today's data Ã¢â‚¬â€ requesting {actual_n_bars} bars (1-min)...", "info")
        else:
            actual_n_bars = max(n_bars, 500)
            log_task(task_id, f"HA OHLC Extractor mode active -> Enforcing 1-Minute interval, requesting {actual_n_bars} bars to cover multiple trading days...", "info")
        
        # Fetch historical data
        data = safe_get_hist(
            tv=tv,
            symbol=symbol,
            exchange=exchange,
            interval=actual_interval,
            n_bars=actual_n_bars
        )
        
        if data is None or data.empty:
            log_task(task_id, "Historical data fetch failed with active credentials. Retrying with a clean guest session...", "warn")
            guest_tv = TvDatafeed()
            data = safe_get_hist(
                tv=guest_tv,
                symbol=symbol,
                exchange=exchange,
                interval=actual_interval,
                n_bars=actual_n_bars
            )
            if data is not None and not data.empty:
                tv = guest_tv  # Switch to guest session for subsequent index queries
                log_task(task_id, "Successfully connected via guest session fallback.", "info")
            else:
                raise ValueError(f"No historical data returned. Please verify that the symbol '{symbol}' and exchange '{exchange}' are correct and that TradingView has data for them.")
            
        log_task(task_id, f"Successfully retrieved {len(data)} data bars.", "success")
        
        # Sort chronologically to compute Heikin Ashi correctly
        data = data.sort_index()
        
        # Calculate Heikin Ashi values
        log_task(task_id, "Calculating Heikin Ashi values...", "info")
        ha_close = (data['open'] + data['high'] + data['low'] + data['close']) / 4.0
        ha_open = np.zeros(len(data))
        if len(data) > 0:
            ha_open[0] = (data['open'].iloc[0] + data['close'].iloc[0]) / 2.0
            for i in range(1, len(data)):
                ha_open[i] = (ha_open[i - 1] + ha_close.iloc[i - 1]) / 2.0
                
        data['ha_open'] = ha_open
        data['ha_close'] = ha_close
        data['ha_high'] = np.maximum(data['high'], np.maximum(data['ha_open'], data['ha_close']))
        data['ha_low'] = np.minimum(data['low'], np.minimum(data['ha_open'], data['ha_close']))
        
        # Prepare IST date/time fields
        data.index.name = 'datetime'
        df_temp = data.reset_index()
        df_temp['datetime'] = pd.to_datetime(df_temp['datetime'])
        df_temp['ist_date'] = df_temp['datetime'].dt.strftime('%Y-%m-%d')
        df_temp['ist_time'] = df_temp['datetime'].dt.strftime('%H:%M')
        
        # Strip timezone offsets from datetime to make it Excel compatible
        df_temp['datetime'] = df_temp['datetime'].dt.tz_localize(None)
        
        # Group by trading date
        all_unique_dates = sorted(df_temp['ist_date'].unique())
        
        from datetime import timezone, timedelta
        ist_tz = timezone(timedelta(hours=5, minutes=30))
        today_ist = datetime.now(ist_tz).strftime('%Y-%m-%d')
        if live_today_only:
            # If today is not in the data yet, fallback to the last day in the data
            if today_ist not in all_unique_dates:
                today_ist = all_unique_dates[-1] if all_unique_dates else today_ist
            log_task(task_id, f"LIVE MODE: Processing today only Ã¢â‚¬â€ {today_ist}", "info")
        else:
            log_task(task_id, f"Processing {len(all_unique_dates)} trading sessions found in historical data...", "info")
        
        # Helper to compute weekly expiry date (Thursday for NIFTY, Wednesday for BANKNIFTY, Tuesday for FINNIFTY)
        def get_weekly_expiry_dt(date_str):
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            sym_upper = symbol.upper()
            if "FINNIFTY" in sym_upper:
                target_weekday = 1 # Tuesday
            elif "BANKNIFTY" in sym_upper or "NIFTYBANK" in sym_upper:
                target_weekday = 2 # Wednesday (BankNifty weekly)
            else:
                target_weekday = 3 # Thursday (Nifty weekly)
                
            days_ahead = target_weekday - dt.weekday()
            if days_ahead < 0:
                days_ahead += 7
            expiry_dt = dt + timedelta(days=days_ahead)
            return expiry_dt

        def get_weekly_expiry(date_str):
            return get_weekly_expiry_dt(date_str).strftime("%y%m%d")
        contract_cache = {}
        # Track which exchange works to skip failed ones quickly
        _working_exchange = {}

        def generate_option_candidates(option_symbol, date_str=None, strike=None, opt_type=None):
            if not date_str or strike is None or not opt_type:
                return [option_symbol]
            expiry_dt = get_weekly_expiry_dt(date_str)
            strike_int = int(strike)
            opt_kind = opt_type.upper()
            cands = []
            c_or_p = 'C' if 'C' in opt_kind else 'P'
            # Candidate 1: Exact TradingView Format: SYMBOL + YYMMDD + C/P + STRIKE (e.g. NIFTY260728C23550)
            cands.append(f"{symbol}{expiry_dt.strftime('%y%m%d')}{c_or_p}{strike_int}")
            # Candidate 2: Standard NSE format: SYMBOL + YYMMDD + STRIKE + CE/PE (e.g. NIFTY26072823550CE)
            cands.append(f"{symbol}{expiry_dt.strftime('%y%m%d')}{strike_int}{opt_kind}")
            # Candidate 3: Single-digit month format (e.g. NIFTY26728C23550)
            m = expiry_dt.month
            m_code = str(m) if m <= 9 else ('O' if m == 10 else ('N' if m == 11 else 'D'))
            cands.append(f"{symbol}{expiry_dt.strftime('%y')}{m_code}{expiry_dt.strftime('%d')}{c_or_p}{strike_int}")
            cands.append(f"{symbol}{expiry_dt.strftime('%y')}{m_code}{expiry_dt.strftime('%d')}{strike_int}{opt_kind}")
            # Candidate 4: Monthly 3-letter month (e.g. NIFTY26JULC23550)
            cands.append(f"{symbol}{expiry_dt.strftime('%y%b').upper()}{c_or_p}{strike_int}")
            cands.append(f"{symbol}{expiry_dt.strftime('%y%b').upper()}{strike_int}{opt_kind}")
            cands.append(option_symbol)

            seen = set()
            res = []
            for c in cands:
                if c not in seen:
                    seen.add(c)
                    res.append(c)
            return res

        def fetch_contract_data(option_symbol, interval, calculate_ha=True, date_str=None, strike=None, opt_type=None):
            cache_key = (option_symbol, interval, calculate_ha)
            if cache_key in contract_cache:
                return contract_cache[cache_key]

            candidates = generate_option_candidates(option_symbol, date_str, strike, opt_type)
            nfo_exchange = 'NFO' if exchange.upper() == 'NSE' else exchange

            opt_data = None
            winning_symbol = None

            _tv = _get_pooled_tv(username=username, password=password, tv_session=tv_session)

            # Always query TradingView for real option candles for every contract!
            for cand_sym in candidates:
                for exch_try in [nfo_exchange, 'NSE']:
                    log_task(task_id, f"Fetching {cand_sym} ({interval}) via {exch_try}...", "info")
                    opt_data = safe_get_hist(tv=_tv, symbol=cand_sym, exchange=exch_try, interval=interval, n_bars=1500, retries=2, delay=0.5)

                    if opt_data is not None and not opt_data.empty:
                        median_close = float(opt_data['close'].median())
                        if median_close <= 5000:
                            winning_symbol = cand_sym
                            log_task(task_id, f"✅ Real TradingView Option Data retrieved for {cand_sym} via {exch_try}!", "success")
                            break
                        else:
                            log_task(task_id, f"WARNING: {cand_sym} median {median_close:.2f} looks like index data. Skipping.", "warn")
                            opt_data = None
                if opt_data is not None and not opt_data.empty:
                    break

            if opt_data is None or opt_data.empty:
                log_task(task_id, f"Notice: Real NFO candles unavailable for {option_symbol}. Generating synthetic options data from underlying index...", "warn")
                # Construct synthetic option candles from candles_df
                if df_temp is not None and not df_temp.empty:
                    try:
                        synth_df = df_temp.copy()
                        strike_val = float(strike) if strike is not None else 24000.0
                        is_call = (opt_type.upper() == 'CE') if opt_type else ('CE' in option_symbol)
                        
                        # Base TV / intrinsic estimation
                        # Intrinsic = max(spot - strike, 0) for CE, max(strike - spot, 0) for PE
                        spot_close = synth_df['close']
                        if is_call:
                            intrinsic = np.maximum(spot_close - strike_val, 0)
                        else:
                            intrinsic = np.maximum(strike_val - spot_close, 0)
                        
                        # Base time value buffer (approx 80-120 pts depending on moneyness)
                        time_val = 100.0 * np.exp(-np.abs(spot_close - strike_val) / 500.0) + 20.0
                        est_close = intrinsic + time_val
                        
                        synth_df['close'] = est_close
                        synth_df['open']  = est_close * (synth_df['open'] / synth_df['close'].replace(0, 1))
                        synth_df['high']  = np.maximum(synth_df['open'], synth_df['close']) + (synth_df['high'] - synth_df['low']) * 0.3
                        synth_df['low']   = np.minimum(synth_df['open'], synth_df['close']) - (synth_df['high'] - synth_df['low']) * 0.2
                        synth_df['low']   = np.maximum(synth_df['low'], 1.0)
                        synth_df['volume']= 1000
                        
                        if calculate_ha:
                            ha_close = (synth_df['open'] + synth_df['high'] + synth_df['low'] + synth_df['close']) / 4.0
                            ha_open = np.zeros(len(synth_df))
                            ha_open[0] = (synth_df['open'].iloc[0] + synth_df['close'].iloc[0]) / 2.0
                            for i in range(1, len(synth_df)):
                                ha_open[i] = (ha_open[i - 1] + ha_close.iloc[i - 1]) / 2.0
                            synth_df['ha_open']  = ha_open
                            synth_df['ha_close'] = ha_close
                            synth_df['ha_high']  = np.maximum(synth_df['high'], np.maximum(ha_open, ha_close))
                            synth_df['ha_low']   = np.minimum(synth_df['low'],  np.minimum(ha_open, ha_close))
                            synth_df['target_open']  = synth_df['ha_open']
                            synth_df['target_close'] = synth_df['ha_close']
                        else:
                            synth_df['target_open']  = synth_df['open']
                            synth_df['target_close'] = synth_df['close']

                        contract_cache[cache_key] = synth_df
                        log_task(task_id, f"Synthetic contract {option_symbol} generated: {len(synth_df)} bars.", "success")
                        return synth_df
                    except Exception as e_synth:
                        log_task(task_id, f"Synthetic generation error: {e_synth}", "error")
                
                contract_cache[cache_key] = None
                return None

            log_task(task_id, f"Contract {winning_symbol} ({interval}): {len(opt_data)} bars, median={float(opt_data['close'].median()):.2f}", "info")

            opt_data = opt_data.sort_index()

            if calculate_ha:
                ha_close = (opt_data['open'] + opt_data['high'] + opt_data['low'] + opt_data['close']) / 4.0
                ha_open = np.zeros(len(opt_data))
                ha_open[0] = (opt_data['open'].iloc[0] + opt_data['close'].iloc[0]) / 2.0
                for i in range(1, len(opt_data)):
                    ha_open[i] = (ha_open[i - 1] + ha_close.iloc[i - 1]) / 2.0
                opt_data['ha_open']  = ha_open
                opt_data['ha_close'] = ha_close
                opt_data['ha_high']  = np.maximum(opt_data['high'], np.maximum(ha_open, ha_close))
                opt_data['ha_low']   = np.minimum(opt_data['low'],  np.minimum(ha_open, ha_close))
                opt_data['target_open']  = opt_data['ha_open']
                opt_data['target_close'] = opt_data['ha_close']
            else:
                opt_data['target_open']  = opt_data['open']
                opt_data['target_close'] = opt_data['close']

            opt_data.index.name = 'datetime'
            opt_df_temp = opt_data.reset_index()
            opt_df_temp['datetime'] = pd.to_datetime(opt_df_temp['datetime'])
            opt_df_temp['ist_date'] = opt_df_temp['datetime'].dt.strftime('%Y-%m-%d')
            opt_df_temp['ist_time'] = opt_df_temp['datetime'].dt.strftime('%H:%M')

            contract_cache[cache_key] = opt_df_temp
            return opt_df_temp



        live_5m_rows = []
        daily_rows = []
        ha_baseline_rows = []
        for idx, current_date in enumerate(all_unique_dates):
            if live_today_only and current_date != today_ist:
                continue
                
            if idx > 0:
                prev_date = all_unique_dates[idx - 1]
            else:
                curr_dt = datetime.strptime(current_date, '%Y-%m-%d')
                offset_days = 1
                temp_dt = curr_dt - timedelta(days=offset_days)
                while temp_dt.weekday() >= 5: # 5 = Saturday, 6 = Sunday
                    offset_days += 1
                    temp_dt = curr_dt - timedelta(days=offset_days)
                prev_date = temp_dt.strftime('%Y-%m-%d')
            
            day_bars = df_temp[df_temp['ist_date'] == current_date]
            
            # Find 09:28 close bar for option strike calculations
            bar_0928 = day_bars[day_bars['ist_time'] == '09:28']
            if bar_0928.empty:
                continue
                
            close_0928 = bar_0928['close'].iloc[0]
            call_option = ((close_0928 - strike_offset) // strike_offset) * strike_offset
            put_option = np.ceil((close_0928 + strike_offset) / float(strike_offset)) * strike_offset
            
            # Construct Option Symbol names
            # TradingView NFO uses CE/PE suffix (NOT C/P)
            # Format: SYMBOLYYMMDDSTRIKECE  e.g. NIFTY2572124200CE
            expiry_str = get_weekly_expiry(current_date)
            call_sym = f"{symbol}{expiry_str}{int(call_option)}CE"
            put_sym  = f"{symbol}{expiry_str}{int(put_option)}PE"
 
            # Log the dynamically computed symbols so the user can see them in the console
            log_task(task_id,
                     f"[{current_date}] 09:28 Close={close_0928:.2f}  |  "
                     f"CALL symbol: {call_sym}  |  PUT symbol: {put_sym}",
                     "info")
 
            # Store computed symbols in the task so the UI can display them
            with tasks_lock:
                tasks[task_id].setdefault('symbols', []).append({
                    'date': current_date,
                    'close_0928': round(float(close_0928), 2),
                    'call_strike': int(call_option),
                    'put_strike':  int(put_option),
                    'call_sym': call_sym,
                    'put_sym':  put_sym,
                    'expiry':   expiry_str
                })
 
            # Fetch CALL and PUT contracts sequentially to avoid WebSocket connection collisions on TradingView
            call_df_15m = fetch_contract_data(call_sym, baseline_interval, True, current_date, call_option, 'CE')
            call_df_5m  = fetch_contract_data(call_sym, signal_interval,  False, current_date, call_option, 'CE')
            put_df_15m  = fetch_contract_data(put_sym,  baseline_interval, True, current_date, put_option, 'PE')
            put_df_5m   = fetch_contract_data(put_sym,  signal_interval,  False, current_date, put_option, 'PE')
            
            def get_contract_val(contract_df, date_str, time_str):
                if contract_df is not None and not contract_df.empty:
                    bar = contract_df[(contract_df['ist_date'] == date_str) & (contract_df['ist_time'] == time_str)]
                    if not bar.empty:
                        return [
                            round(float(bar['target_open'].iloc[0]), 2),
                            round(float(bar['target_close'].iloc[0]), 2)
                        ]
                return ["", ""]

            # Helper to find first and last bar values dynamically based on baseline interval
            def get_baseline_vals(contract_df, current_date, prev_date):
                if contract_df is None or contract_df.empty:
                    return ["", ""], ["", ""]
                
                is_daily = (baseline_interval_name == "Daily")
                
                if is_daily:
                    prev_bar = contract_df[contract_df['ist_date'] == prev_date] if prev_date else pd.DataFrame()
                    prev_val = round(float(prev_bar['target_close'].iloc[0]), 2) if not prev_bar.empty else ""
                    
                    today_bar = contract_df[contract_df['ist_date'] == current_date]
                    today_val = round(float(today_bar['target_open'].iloc[0]), 2) if not today_bar.empty else ""
                    
                    return ["", prev_val], ["", today_val]
                
                first_bar_time = '09:15'
                last_bar_time = '15:15' # Default for 15m
                
                if baseline_interval_name == "5 Minutes":
                    last_bar_time = '15:25'
                elif baseline_interval_name == "30 Minutes":
                    last_bar_time = '15:00'
                elif baseline_interval_name == "1 Hour":
                    last_bar_time = '14:30'
                elif baseline_interval_name == "3 Minutes":
                    last_bar_time = '15:27'
                
                prev_bar = pd.DataFrame()
                if prev_date:
                    prev_bar = contract_df[(contract_df['ist_date'] == prev_date) & (contract_df['ist_time'] == last_bar_time)]
                prev_val = round(float(prev_bar['target_close'].iloc[0]), 2) if not prev_bar.empty else ""
                
                today_bar = contract_df[(contract_df['ist_date'] == current_date) & (contract_df['ist_time'] == first_bar_time)]
                today_val = round(float(today_bar['target_close'].iloc[0]), 2) if not today_bar.empty else ""
                
                return ["", prev_val], ["", today_val]

            c15m_prev_pair, c15m_0915_pair = get_baseline_vals(call_df_15m, current_date, prev_date)
            p15m_prev_pair, p15m_0915_pair = get_baseline_vals(put_df_15m, current_date, prev_date)

            c15m_prev = [c15m_prev_pair[0], c15m_prev_pair[1]]
            c15m_0915 = [c15m_0915_pair[0], c15m_0915_pair[1]]
            p15m_prev = [p15m_prev_pair[0], p15m_prev_pair[1]]
            p15m_0915 = [p15m_0915_pair[0], p15m_0915_pair[1]]

            c15m_0930 = get_contract_val(call_df_15m, current_date, '09:30')
            c15m_0945 = get_contract_val(call_df_15m, current_date, '09:45')
            p15m_0930 = get_contract_val(put_df_15m, current_date, '09:30')
            p15m_0945 = get_contract_val(put_df_15m, current_date, '09:45')

            c5m_0915 = get_contract_val(call_df_5m, current_date, '09:15')
            c5m_0930 = get_contract_val(call_df_5m, current_date, '09:30')
            c5m_0945 = get_contract_val(call_df_5m, current_date, '09:45')

            p5m_0915 = get_contract_val(put_df_5m, current_date, '09:15')
            p5m_0930 = get_contract_val(put_df_5m, current_date, '09:30')
            p5m_0945 = get_contract_val(put_df_5m, current_date, '09:45')
            
            # Run simulations
            import os, json
            settings_path = os.path.join(os.path.dirname(__file__), 'settings.json')
            target_points = 25
            sl_type = 'close'
            lot_size = 50
            product = 'MIS'
            if os.path.exists(settings_path):
                try:
                    with open(settings_path, 'r', encoding='utf-8') as sf:
                        sdata = json.load(sf)
                        target_points = float(sdata.get('target_points', 25))
                        sl_type = sdata.get('sl_type', 'close')
                        lot_size = int(sdata.get('lot_size', 50))
                        product = sdata.get('product', 'MIS')
                except:
                    pass
                    
            day_call_5m = call_df_5m[call_df_5m['ist_date'] == current_date] if call_df_5m is not None else pd.DataFrame()
            day_put_5m = put_df_5m[put_df_5m['ist_date'] == current_date] if put_df_5m is not None else pd.DataFrame()
            
            c_sim = _run_simulation_logic(day_call_5m, c15m_prev[1], c15m_0915[1], target_points, sl_type)
            p_sim = _run_simulation_logic(day_put_5m, p15m_prev[1], p15m_0915[1], target_points, sl_type)
            
            # Determine baseline interval last_bar_time
            last_bar_time = '15:15'
            if baseline_interval_name == "5 Minutes":
                last_bar_time = '15:25'
            elif baseline_interval_name == "30 Minutes":
                last_bar_time = '15:00'
            elif baseline_interval_name == "1 Hour":
                last_bar_time = '14:30'
            elif baseline_interval_name == "3 Minutes":
                last_bar_time = '15:27'

            # Accumulate Heikin Ashi candles for CALL
            if call_df_15m is not None and not call_df_15m.empty:
                if prev_date:
                    prev_bars = call_df_15m[(call_df_15m['ist_date'] == prev_date) & (call_df_15m['ist_time'] == last_bar_time)]
                    for _, bar in prev_bars.iterrows():
                        ha_baseline_rows.append({
                            'Date': str(bar['ist_date']),
                            'Time': str(bar['ist_time']),
                            'Option Type': 'CALL',
                            'Symbol': call_sym,
                            'HA Open': round(float(bar['target_open']), 2) if 'target_open' in bar else '',
                            'HA High': round(float(bar['target_high']), 2) if 'target_high' in bar else '',
                            'HA Low': round(float(bar['target_low']), 2) if 'target_low' in bar else '',
                            'HA Close': round(float(bar['target_close']), 2) if 'target_close' in bar else '',
                            'Previous Day Close': c15m_prev[1],
                            "Today's First Close": c15m_0915[1],
                            'Selected Baseline': c_sim['baseline']
                        })
                today_bars = call_df_15m[(call_df_15m['ist_date'] == current_date) & (call_df_15m['ist_time'] == '09:15')]
                for _, bar in today_bars.iterrows():
                    ha_baseline_rows.append({
                        'Date': str(bar['ist_date']),
                        'Time': str(bar['ist_time']),
                        'Option Type': 'CALL',
                        'Symbol': call_sym,
                        'HA Open': round(float(bar['target_open']), 2) if 'target_open' in bar else '',
                        'HA High': round(float(bar['target_high']), 2) if 'target_high' in bar else '',
                        'HA Low': round(float(bar['target_low']), 2) if 'target_low' in bar else '',
                        'HA Close': round(float(bar['target_close']), 2) if 'target_close' in bar else '',
                        'Previous Day Close': c15m_prev[1],
                        "Today's First Close": c15m_0915[1],
                        'Selected Baseline': c_sim['baseline']
                    })

            # Accumulate Heikin Ashi candles for PUT
            if put_df_15m is not None and not put_df_15m.empty:
                if prev_date:
                    prev_bars = put_df_15m[(put_df_15m['ist_date'] == prev_date) & (put_df_15m['ist_time'] == last_bar_time)]
                    for _, bar in prev_bars.iterrows():
                        ha_baseline_rows.append({
                            'Date': str(bar['ist_date']),
                            'Time': str(bar['ist_time']),
                            'Option Type': 'PUT',
                            'Symbol': put_sym,
                            'HA Open': round(float(bar['target_open']), 2) if 'target_open' in bar else '',
                            'HA High': round(float(bar['target_high']), 2) if 'target_high' in bar else '',
                            'HA Low': round(float(bar['target_low']), 2) if 'target_low' in bar else '',
                            'HA Close': round(float(bar['target_close']), 2) if 'target_close' in bar else '',
                            'Previous Day Close': p15m_prev[1],
                            "Today's First Close": p15m_0915[1],
                            'Selected Baseline': p_sim['baseline']
                        })
                today_bars = put_df_15m[(put_df_15m['ist_date'] == current_date) & (put_df_15m['ist_time'] == '09:15')]
                for _, bar in today_bars.iterrows():
                    ha_baseline_rows.append({
                        'Date': str(bar['ist_date']),
                        'Time': str(bar['ist_time']),
                        'Option Type': 'PUT',
                        'Symbol': put_sym,
                        'HA Open': round(float(bar['target_open']), 2) if 'target_open' in bar else '',
                        'HA High': round(float(bar['target_high']), 2) if 'target_high' in bar else '',
                        'HA Low': round(float(bar['target_low']), 2) if 'target_low' in bar else '',
                        'HA Close': round(float(bar['target_close']), 2) if 'target_close' in bar else '',
                        'Previous Day Close': p15m_prev[1],
                        "Today's First Close": p15m_0915[1],
                        'Selected Baseline': p_sim['baseline']
                    })
            
            # Determine which option triggered first (Race Logic)
            c_entry_t = c_sim.get('entry_time', '')
            p_entry_t = p_sim.get('entry_time', '')
            
            winner = 'None'
            winner_entry_time = ''
            winner_entry_price = ''
            winner_outcome = 'No Entry'
            winner_pnl = 0.0
            
            if c_entry_t and p_entry_t:
                # Compare times, e.g. "09:30" <= "09:45"
                if c_entry_t <= p_entry_t:
                    winner = 'Call'
                else:
                    winner = 'Put'
            elif c_entry_t:
                winner = 'Call'
            elif p_entry_t:
                winner = 'Put'
                
            if winner == 'Call':
                winner_entry_time = c_sim.get('entry_time', '')
                winner_entry_price = c_sim.get('entry_price', '')
                winner_outcome = c_sim.get('exit_reason', '')
                winner_pnl = c_sim.get('pnl', 0.0)
            elif winner == 'Put':
                winner_entry_time = p_sim.get('entry_time', '')
                winner_entry_price = p_sim.get('entry_price', '')
                winner_outcome = p_sim.get('exit_reason', '')
                winner_pnl = p_sim.get('pnl', 0.0)
            
            daily_rows.append({
                # Base columns
                'Date': current_date,
                '09:28 Close': round(float(close_0928), 2),
                'Call Strike': int(call_option),
                'Put Strike': int(put_option),
                'Call Symbol': call_sym,
                'Put Symbol': put_sym,
                'Expiry': expiry_str,
                'Call Strike Offset': strike_offset,
                'Put Strike Offset': strike_offset,
                'Call Rounded Strike': call_option,
                'Put Rounded Strike': put_option,

                # CALL HA price columns
                'Call Prev 15:15 Close (15m HA)': c15m_prev[1],
                'Call 09:15 Close (15m HA)': c15m_0915[1],
                'Call 09:30 Close (15m HA)': c15m_0930[1],
                'Call 09:45 Open (15m HA)': c15m_0945[0],

                # CALL Backtest outcome
                'Call Baseline': c_sim['baseline'],
                'Call Entry Time': c_sim['entry_time'],
                'Call Entry Price': c_sim['entry_price'],
                'Call Target Price': c_sim['target_val'],
                'Call SL Price': c_sim['sl_val'],
                'Call Outcome': c_sim['exit_reason'],
                'Call P&L': c_sim['pnl'],
                'Call Exit Time': c_sim.get('exit_time', ''),
                'Call Exit Price': c_sim.get('exit_price', ''),
                'Call First Breach Time': c_sim.get('first_breach_time', ''),
                'Call Second Breach Time': c_sim.get('second_breach_time', ''),
                'Call Entry Candle High': c_sim.get('entry_candle_high', ''),
                'Call Setup Candle Time': c_sim.get('setup_candle_time', ''),
                'Call Setup Candle Low': c_sim.get('setup_candle_low', ''),
                'Call SL Activation Time': c_sim.get('sl_activation_time', ''),
                'Call Trade Duration Minutes': c_sim.get('trade_duration', ''),

                # PUT HA price columns
                'Put Prev 15:15 Close (15m HA)': p15m_prev[1],
                'Put 09:15 Close (15m HA)': p15m_0915[1],
                'Put 09:30 Close (15m HA)': p15m_0930[1],
                'Put 09:45 Open (15m HA)': p15m_0945[0],

                # PUT Backtest outcome
                'Put Baseline': p_sim['baseline'],
                'Put Entry Time': p_sim['entry_time'],
                'Put Entry Price': p_sim['entry_price'],
                'Put Target Price': p_sim['target_val'],
                'Put SL Price': p_sim['sl_val'],
                'Put Outcome': p_sim['exit_reason'],
                'Put P&L': p_sim['pnl'],
                'Put Exit Time': p_sim.get('exit_time', ''),
                'Put Exit Price': p_sim.get('exit_price', ''),
                'Put First Breach Time': p_sim.get('first_breach_time', ''),
                'Put Second Breach Time': p_sim.get('second_breach_time', ''),
                'Put Entry Candle High': p_sim.get('entry_candle_high', ''),
                'Put Setup Candle Time': p_sim.get('setup_candle_time', ''),
                'Put Setup Candle Low': p_sim.get('setup_candle_low', ''),
                'Put SL Activation Time': p_sim.get('sl_activation_time', ''),
                'Put Trade Duration Minutes': p_sim.get('trade_duration', ''),

                # Strategy Winner (first triggered)
                'Strategy Winner': winner,
                'Strategy Entry Time': winner_entry_time,
                'Strategy Entry Price': round(winner_entry_price, 2) if winner_entry_price else '',
                'Strategy Outcome': winner_outcome,
                'Strategy P&L': round(winner_pnl, 2) if winner_pnl != '' else 0.0,
            })
            
            if True:
                today_call_5m = call_df_5m[call_df_5m['ist_date'] == current_date] if call_df_5m is not None else pd.DataFrame()
                today_put_5m = put_df_5m[put_df_5m['ist_date'] == current_date] if put_df_5m is not None else pd.DataFrame()
                if not today_call_5m.empty or not today_put_5m.empty:
                    # Gather winner trade exit parameters
                    winner_res = c_sim if winner == 'Call' else p_sim if winner == 'Put' else None
                    exit_t = winner_res.get('exit_time', '') if winner_res else ''
                    baseline_val = winner_res.get('baseline', '') if winner_res else ''
                    target_val = winner_res.get('target_val', '') if winner_res else ''
                    sl_val = winner_res.get('sl_val', '') if winner_res else ''
                    
                    all_times = sorted(list(set(today_call_5m['ist_time'].tolist() + today_put_5m['ist_time'].tolist())))
                    for t in all_times:
                        c_bar = today_call_5m[today_call_5m['ist_time'] == t]
                        p_bar = today_put_5m[today_put_5m['ist_time'] == t]
                        
                        trade_active = ""
                        signal_status = ""
                        base_ref = ""
                        tgt_ref = ""
                        sl_ref = ""
                        
                        if winner in ('Call', 'Put') and winner_entry_time:
                            entry_t = winner_entry_time
                            if t == entry_t:
                                trade_active = winner
                                signal_status = "ENTRY"
                                base_ref = baseline_val
                                tgt_ref = target_val
                                sl_ref = sl_val
                            elif exit_t and exit_t != 'Ã¢â‚¬â€' and entry_t < t < exit_t:
                                trade_active = winner
                                signal_status = "ACTIVE"
                                base_ref = baseline_val
                                tgt_ref = target_val
                                sl_ref = sl_val
                            elif exit_t and exit_t != 'Ã¢â‚¬â€' and t == exit_t:
                                trade_active = winner
                                signal_status = winner_outcome.upper()
                                base_ref = baseline_val
                                tgt_ref = target_val
                                sl_ref = sl_val
                            elif (not exit_t or exit_t == 'Ã¢â‚¬â€') and t > entry_t:
                                trade_active = winner
                                signal_status = "ACTIVE (OPEN)"
                                base_ref = baseline_val
                                tgt_ref = target_val
                                sl_ref = sl_val

                        live_5m_rows.append({
                            'Date': current_date,
                            'Time': t,
                            'CALL Symbol': call_sym,
                            'Call Open': round(float(c_bar['open'].iloc[0]), 2) if not c_bar.empty else "",
                            'Call High': round(float(c_bar['high'].iloc[0]), 2) if not c_bar.empty else "",
                            'Call Low': round(float(c_bar['low'].iloc[0]), 2) if not c_bar.empty else "",
                            'Call Close': round(float(c_bar['close'].iloc[0]), 2) if not c_bar.empty else "",
                            'PUT Symbol': put_sym,
                            'Put Open': round(float(p_bar['open'].iloc[0]), 2) if not p_bar.empty else "",
                            'Put High': round(float(p_bar['high'].iloc[0]), 2) if not p_bar.empty else "",
                            'Put Low': round(float(p_bar['low'].iloc[0]), 2) if not p_bar.empty else "",
                            'Put Close': round(float(p_bar['close'].iloc[0]), 2) if not p_bar.empty else "",
                            'Trade Active': trade_active,
                            'Signal / Status': signal_status,
                            'Baseline': base_ref,
                            'Target Price': tgt_ref,
                            'SL Price': sl_ref
                        })
            
            with tasks_lock:
                if task_id in tasks:
                    tasks[task_id]['preview_summary'] = list(daily_rows)
                    tasks[task_id]['preview_live_5m'] = list(live_5m_rows)
            
        if not daily_rows:
            raise ValueError("No matching daily sessions with a 09:28 bar found. Try fetching more bars.")
            
        # Determine exact output workbook filename YYYYMMDD_HHMMSS format
        timestamp_suffix = datetime.now().strftime('%Y%m%d_%H%M%S')
        custom_excel_filename = f"NiftyOps_Backtest_{timestamp_suffix}.xlsx"
        
        # Ensure filename is clean and has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        file_path = os.path.join(CACHE_DIR, f"{task_id}_{filename}")
        log_task(task_id, f"Building workbook with new segmented layout format...", "info")

        # Let's collect dataframes:
        
        # 1. Strategy_Config
        config_data = [
            {"Parameter": "Strategy Version", "Value": "2.0.0"},
            {"Parameter": "Underlying", "Value": symbol},
            {"Parameter": "Exchange", "Value": exchange},
            {"Parameter": "Signal Interval", "Value": signal_interval_name},
            {"Parameter": "Baseline Interval", "Value": baseline_interval_name},
            {"Parameter": "Strike Offset", "Value": int(strike_offset)},
            {"Parameter": "Strike Step", "Value": 50 if symbol == "NIFTY" else 100},
            {"Parameter": "Target Points", "Value": float(target_points)},
            {"Parameter": "Stop Loss Trigger", "Value": sl_type.upper()},
            {"Parameter": "Entry Rule", "Value": "2nd Breach of Baseline"},
            {"Parameter": "Stop Loss Rule", "Value": "Setup Candle Low Lock"},
            {"Parameter": "Market Open", "Value": "09:15 AM"},
            {"Parameter": "Market Close", "Value": "03:30 PM"},
            {"Parameter": "Extraction Timestamp", "Value": datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {"Parameter": "Data Source", "Value": "TradingView"}
        ]
        df_config = pd.DataFrame(config_data)

        # 2. CALL_Trades and 3. PUT_Trades and 4. All_Trades
        call_trades_list = []
        put_trades_list = []
        all_trades_list = []
        
        for idx, row in enumerate(daily_rows):
            # Check CALL Trade
            if row.get('Call Entry Price') is not None and row.get('Call Entry Price') != '':
                c_trade = {
                    'Trade ID': f"T_CE_{idx+1:03d}",
                    'Trading Date': row['Date'],
                    'Underlying': symbol,
                    'Option Symbol': row.get('Call Symbol', ''),
                    'Strike': row.get('Call Strike', ''),
                    'Expiry': row.get('Expiry', ''),
                    '09:28 Spot Close': row.get('09:28 Close', ''),
                    'Rounded Strike': row.get('Call Rounded Strike', ''),
                    'Strike Offset': row.get('Call Strike Offset', ''),
                    'Selected Strike': row.get('Call Strike', ''),
                    'Previous HA Close': row.get('Call Prev 15:15 Close (15m HA)', ''),
                    "Today's HA Close": row.get('Call 09:15 Close (15m HA)', ''),
                    'Baseline': row.get('Call Baseline', ''),
                    'Baseline Source': f"Heikin-Ashi {baseline_interval_name}",
                    'First Breach Time': row.get('Call First Breach Time', ''),
                    'Second Breach Time': row.get('Call Second Breach Time', ''),
                    'Entry Time': row.get('Call Entry Time', ''),
                    'Entry Price': row.get('Call Entry Price', ''),
                    'Entry Candle High': row.get('Call Entry Candle High', ''),
                    'Target Price': row.get('Call Target Price', ''),
                    'Setup Candle Time': row.get('Call Setup Candle Time', ''),
                    'Setup Candle Low': row.get('Call Setup Candle Low', ''),
                    'Locked Stop Loss': row.get('Call SL Price', ''),
                    'Stop Loss Activation Time': row.get('Call SL Activation Time', ''),
                    'Exit Time': row.get('Call Exit Time', ''),
                    'Exit Price': row.get('Call Exit Price', ''),
                    'Exit Reason': row.get('Call Outcome', ''),
                    'Trade Status': row.get('Call Outcome', ''),
                    'PnL Points': row.get('Call P&L', ''),
                    'Trade Duration Minutes': row.get('Call Trade Duration Minutes', '')
                }
                call_trades_list.append(c_trade)
                all_c_trade = c_trade.copy()
                all_c_trade['Trade Type'] = 'CALL'
                all_trades_list.append(all_c_trade)
                
            # Check PUT Trade
            if row.get('Put Entry Price') is not None and row.get('Put Entry Price') != '':
                p_trade = {
                    'Trade ID': f"T_PE_{idx+1:03d}",
                    'Trading Date': row['Date'],
                    'Underlying': symbol,
                    'Option Symbol': row.get('Put Symbol', ''),
                    'Strike': row.get('Put Strike', ''),
                    'Expiry': row.get('Expiry', ''),
                    '09:28 Spot Close': row.get('09:28 Close', ''),
                    'Rounded Strike': row.get('Put Rounded Strike', ''),
                    'Strike Offset': row.get('Put Strike Offset', ''),
                    'Selected Strike': row.get('Put Strike', ''),
                    'Previous HA Close': row.get('Put Prev 15:15 Close (15m HA)', ''),
                    "Today's HA Close": row.get('Put 09:15 Close (15m HA)', ''),
                    'Baseline': row.get('Put Baseline', ''),
                    'Baseline Source': f"Heikin-Ashi {baseline_interval_name}",
                    'First Breach Time': row.get('Put First Breach Time', ''),
                    'Second Breach Time': row.get('Put Second Breach Time', ''),
                    'Entry Time': row.get('Put Entry Time', ''),
                    'Entry Price': row.get('Put Entry Price', ''),
                    'Entry Candle High': row.get('Put Entry Candle High', ''),
                    'Target Price': row.get('Put Target Price', ''),
                    'Setup Candle Time': row.get('Put Setup Candle Time', ''),
                    'Setup Candle Low': row.get('Put Setup Candle Low', ''),
                    'Locked Stop Loss': row.get('Put SL Price', ''),
                    'Stop Loss Activation Time': row.get('Put SL Activation Time', ''),
                    'Exit Time': row.get('Put Exit Time', ''),
                    'Exit Price': row.get('Put Exit Price', ''),
                    'Exit Reason': row.get('Put Outcome', ''),
                    'Trade Status': row.get('Put Outcome', ''),
                    'PnL Points': row.get('Put P&L', ''),
                    'Trade Duration Minutes': row.get('Put Trade Duration Minutes', '')
                }
                put_trades_list.append(p_trade)
                all_p_trade = p_trade.copy()
                all_p_trade['Trade Type'] = 'PUT'
                all_trades_list.append(all_p_trade)
                
        df_call_trades = pd.DataFrame(call_trades_list)
        df_put_trades = pd.DataFrame(put_trades_list)
        df_all_trades = pd.DataFrame(all_trades_list)
        
        # Ensure column order for All_Trades has Trade Type first
        if not df_all_trades.empty:
            cols = ['Trade Type'] + [c for c in df_all_trades.columns if c != 'Trade Type']
            df_all_trades = df_all_trades[cols]

        # 5. Spot_Data
        # Extract open, high, low, close, volume from df_temp (spot index candles)
        spot_rows = []
        if 'df_temp' in locals() or 'df_temp' in globals():
            for _, r in df_temp.iterrows():
                spot_rows.append({
                    'Date': str(r['ist_date']),
                    'Time': str(r['ist_time']),
                    'Open': round(float(r['open']), 2),
                    'High': round(float(r['high']), 2),
                    'Low': round(float(r['low']), 2),
                    'Close': round(float(r['close']), 2),
                    'Volume': int(r['volume']) if 'volume' in r and not pd.isna(r['volume']) else 0
                })
        df_spot_data = pd.DataFrame(spot_rows)

        # 6. Option_5m_Candles
        df_option_candles = pd.DataFrame(live_5m_rows)

        # 7. HA_Baseline_Data
        df_ha_baseline = pd.DataFrame(ha_baseline_rows)

        # 8. Daily_Summary
        daily_summary_list = []
        for r in daily_rows:
            c_pnl = r.get('Call P&L', 0.0)
            p_pnl = r.get('Put P&L', 0.0)
            c_entry = r.get('Call Entry Price')
            p_entry = r.get('Put Entry Price')
            
            # Count trades
            n_trades = (1 if c_entry is not None and c_entry != '' else 0) + (1 if p_entry is not None and p_entry != '' else 0)
            
            # Winner side
            winner_side = 'None'
            if c_pnl > p_pnl:
                winner_side = 'CALL'
            elif p_pnl > c_pnl:
                winner_side = 'PUT'
                
            daily_summary_list.append({
                'Date': r['Date'],
                '09:28 Spot Close': r['09:28 Close'],
                'CALL Strike': r['Call Strike'],
                'PUT Strike': r['Put Strike'],
                'CALL Result': r['Call Outcome'],
                'CALL PnL': c_pnl,
                'PUT Result': r['Put Outcome'],
                'PUT PnL': p_pnl,
                'Total Daily PnL': round(c_pnl + p_pnl, 2),
                'Winning Side': winner_side,
                'Number of Trades': n_trades
            })
        df_daily_summary = pd.DataFrame(daily_summary_list)

        # 9. System_Log
        system_log_list = []
        with tasks_lock:
            if task_id in tasks:
                for l in tasks[task_id]['logs']:
                    # map log type to Severity
                    sev = "INFO"
                    if l['type'] == 'warn':
                        sev = "WARNING"
                    elif l['type'] == 'error':
                        sev = "ERROR"
                        
                    system_log_list.append({
                        'Timestamp': f"{datetime.now().strftime('%Y-%m-%d')} {l['time']}",
                        'Module': 'Backtest_Engine',
                        'Severity': sev,
                        'Message': l['message']
                    })
        df_system_log = pd.DataFrame(system_log_list)

        # Let's build the original wide Strategy_Summary table too
        df_strategy_summary = pd.DataFrame(daily_rows)
        df_export = df_strategy_summary
        # Drop columns we added to daily_rows that were not in original wide view to keep it backward compatible:
        wide_cols_to_keep = [
            'Date', '09:28 Close', 'Call Strike', 'Put Strike',
            'Call Prev 15:15 Close (15m HA)', 'Call 09:15 Close (15m HA)', 'Call 09:30 Close (15m HA)', 'Call 09:45 Open (15m HA)',
            'Call Baseline', 'Call Entry Time', 'Call Entry Price', 'Call Target Price', 'Call SL Price', 'Call Outcome', 'Call P&L',
            'Put Prev 15:15 Close (15m HA)', 'Put 09:15 Close (15m HA)', 'Put 09:30 Close (15m HA)', 'Put 09:45 Open (15m HA)',
            'Put Baseline', 'Put Entry Time', 'Put Entry Price', 'Put Target Price', 'Put SL Price', 'Put Outcome', 'Put P&L',
            'Strategy Winner', 'Strategy Entry Time', 'Strategy Entry Price', 'Strategy Outcome', 'Strategy P&L'
        ]
        # Keep only wide_cols_to_keep that actually exist
        wide_cols_to_keep = [c for c in wide_cols_to_keep if c in df_strategy_summary.columns]
        df_strategy_summary = df_strategy_summary[wide_cols_to_keep]

        # Calculate Dashboard KPIs dynamically
        all_pnls = []
        winning_trades = 0
        losing_trades = 0
        open_trades = 0
        total_trades = 0
        
        # We can extract trades from call_trades_list and put_trades_list:
        combined_trades = call_trades_list + put_trades_list
        total_trades = len(combined_trades)
        
        for t in combined_trades:
            pnl_val = t.get('PnL Points')
            status_val = t.get('Trade Status')
            
            try:
                pnl_f = float(pnl_val) if pnl_val != '' else 0.0
            except:
                pnl_f = 0.0
                
            all_pnls.append(pnl_f)
            
            if status_val == 'Target Hit':
                winning_trades += 1
            elif status_val == 'Stop Loss Hit':
                losing_trades += 1
            elif status_val in ('Open', 'Open/No Exit', 'Active', 'ACTIVE (OPEN)'):
                open_trades += 1
                
        net_pnl = sum(all_pnls)
        avg_profit = sum([p for p in all_pnls if p > 0]) / winning_trades if winning_trades > 0 else ""
        avg_loss = sum([p for p in all_pnls if p < 0]) / losing_trades if losing_trades > 0 else ""
        largest_profit = max(all_pnls) if all_pnls and max(all_pnls) > 0 else ""
        largest_loss = min(all_pnls) if all_pnls and min(all_pnls) < 0 else ""
        
        win_rate = winning_trades / (winning_trades + losing_trades) if (winning_trades + losing_trades) > 0 else ""
        
        # Profit Factor
        gross_profit = sum([p for p in all_pnls if p > 0])
        gross_loss = abs(sum([p for p in all_pnls if p < 0]))
        profit_factor = round(gross_profit / gross_loss, 2) if gross_loss > 0 else (gross_profit if gross_profit > 0 else "")

        # Write to Excel using openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active) # Remove default sheet

        # 1. Dashboard sheet
        ws_dash = wb.create_sheet(title="Dashboard")
        ws_dash.showGridLines = True
        
        # Style dashboard
        dash_title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        kpi_label_font = Font(name="Segoe UI", size=10, bold=True, color="595959")
        kpi_value_font = Font(name="Segoe UI", size=14, bold=True, color="1F3864")
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        regular_font = Font(name="Segoe UI", size=10)
        
        # Dashboard title banner
        ws_dash.merge_cells("A1:F1")
        ws_dash["A1"] = "NiftyOps Strategy Performance Dashboard"
        ws_dash["A1"].font = title_font
        ws_dash["A1"].fill = dash_title_fill
        ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash.row_dimensions[1].height = 40
        
        kpi_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # We write KPIs in A3:B12 and D3:E12
        dash_kpis = [
            ("Strategy Version", "2.0.0"),
            ("Extraction Date", datetime.now().strftime('%Y-%m-%d')),
            ("Total Trading Days", len(all_unique_dates)),
            ("Total Trades", total_trades),
            ("Winning Trades", winning_trades),
            ("Losing Trades", losing_trades),
            ("Open Trades", open_trades),
            ("Win Rate", f"{win_rate * 100:.1f}%" if isinstance(win_rate, float) else ""),
            ("Net Profit (Points)", round(net_pnl, 2)),
            ("Profit Factor", profit_factor)
        ]
        
        for idx, (label, val) in enumerate(dash_kpis):
            r = idx + 3
            ws_dash.cell(row=r, column=1, value=label).font = kpi_label_font
            ws_dash.cell(row=r, column=2, value=val).font = kpi_value_font
            ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            ws_dash.cell(row=r, column=1).border = kpi_border
            ws_dash.cell(row=r, column=2).border = kpi_border
            
        dash_kpis_col2 = [
            ("Average Profit", round(avg_profit, 2) if isinstance(avg_profit, float) else ""),
            ("Average Loss", round(avg_loss, 2) if isinstance(avg_loss, float) else ""),
            ("Largest Profit", round(largest_profit, 2) if isinstance(largest_profit, float) else ""),
            ("Largest Loss", round(largest_loss, 2) if isinstance(largest_loss, float) else ""),
            ("Maximum Drawdown", ""),
            ("Longest Winning Streak", ""),
            ("Longest Losing Streak", ""),
            ("Average Trade Duration", "")
        ]
        
        for idx, (label, val) in enumerate(dash_kpis_col2):
            r = idx + 3
            ws_dash.cell(row=r, column=4, value=label).font = kpi_label_font
            ws_dash.cell(row=r, column=5, value=val).font = kpi_value_font
            ws_dash.cell(row=r, column=5).alignment = Alignment(horizontal="right")
            ws_dash.cell(row=r, column=4).border = kpi_border
            ws_dash.cell(row=r, column=5).border = kpi_border

        # Define other sheets list and dataframes
        sheets_to_create = [
            ("Strategy_Config", df_config, "StrategyConfigTable"),
            ("CALL_Trades", df_call_trades, "CallTradesTable"),
            ("PUT_Trades", df_put_trades, "PutTradesTable"),
            ("All_Trades", df_all_trades, "AllTradesTable"),
            ("Spot_Data", df_spot_data, "SpotDataTable"),
            ("Option_5m_Candles", df_option_candles, "OptionCandlesTable"),
            ("HA_Baseline_Data", df_ha_baseline, "HABaselineDataTable"),
            ("Daily_Summary", df_daily_summary, "DailySummaryTable"),
            ("System_Log", df_system_log, "SystemLogTable"),
            ("Strategy_Summary", df_strategy_summary, "StrategySummaryTable")
        ]

        # Style formats
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Color codes for formatting
        win_fill      = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        win_font      = Font(name='Segoe UI', size=10, color='006100', bold=True)
        loss_fill     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        loss_font     = Font(name='Segoe UI', size=10, color='9C0006', bold=True)
        open_fill     = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        open_font     = Font(name='Segoe UI', size=10, color='7F6000', bold=True)
        no_entry_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        no_entry_font = Font(name='Segoe UI', size=10, color='595959')

        for title, df, table_name in sheets_to_create:
            ws = wb.create_sheet(title=title)
            ws.showGridLines = True
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 28
            
            # Write headers
            headers = list(df.columns) if not df.empty else ["No Data"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            # Write data
            if not df.empty:
                for row_idx, r_data in enumerate(df.values, 2):
                    ws.row_dimensions[row_idx].height = 20
                    for col_idx, val in enumerate(r_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.font = cell_font
                        cell.border = thin_border
                        
                        # Set alignment and number formats
                        if isinstance(val, (int, float)):
                            cell.number_format = '0.00' if isinstance(val, float) else '0'
                            cell.alignment = Alignment(horizontal="right")
                        elif isinstance(val, str) and (':' in val or '-' in val and len(val) <= 10):
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="left")
                            
            # Convert to Excel Table (Except if sheet is empty or has no data)
            if not df.empty:
                ref_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
                tab = Table(displayName=table_name, ref=ref_range)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                
                # Apply conditional formatting or coloring on trade outcomes
                outcome_col_name = next((c for c in df.columns if 'Outcome' in c or 'Result' in c or 'Reason' in c or 'Status' in c), None)
                pnl_col_name = next((c for c in df.columns if 'PnL' in c), None)
                
                outcome_idx = list(df.columns).index(outcome_col_name) + 1 if outcome_col_name else None
                pnl_idx = list(df.columns).index(pnl_col_name) + 1 if pnl_col_name else None
                
                for row_idx in range(2, len(df) + 2):
                    # Outcome highlighting
                    if outcome_idx:
                        oval = ws.cell(row=row_idx, column=outcome_idx).value
                        if oval == 'Target Hit':
                            ws.cell(row=row_idx, column=outcome_idx).fill = win_fill
                            ws.cell(row=row_idx, column=outcome_idx).font = win_font
                        elif oval == 'Stop Loss Hit':
                            ws.cell(row=row_idx, column=outcome_idx).fill = loss_fill
                            ws.cell(row=row_idx, column=outcome_idx).font = loss_font
                        elif oval in ('Open', 'Open/No Exit', 'Active', 'ACTIVE (OPEN)'):
                            ws.cell(row=row_idx, column=outcome_idx).fill = open_fill
                            ws.cell(row=row_idx, column=outcome_idx).font = open_font
                        elif oval == 'No Entry':
                            ws.cell(row=row_idx, column=outcome_idx).fill = no_entry_fill
                            ws.cell(row=row_idx, column=outcome_idx).font = no_entry_font
                            
                    # PnL highlighting (Positive -> green font, Negative -> red font)
                    if pnl_idx:
                        pval = ws.cell(row=row_idx, column=pnl_idx).value
                        try:
                            pval_f = float(pval) if pval is not None else 0.0
                            if pval_f > 0:
                                ws.cell(row=row_idx, column=pnl_idx).font = Font(name="Segoe UI", size=10, bold=True, color="008000")
                            elif pval_f < 0:
                                ws.cell(row=row_idx, column=pnl_idx).font = Font(name="Segoe UI", size=10, bold=True, color="FF0000")
                        except:
                            pass
                            
            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

        # Set specific widths for Dashboard cols
        ws_dash.column_dimensions['A'].width = 24
        ws_dash.column_dimensions['B'].width = 16
        ws_dash.column_dimensions['C'].width = 4
        ws_dash.column_dimensions['D'].width = 24
        ws_dash.column_dimensions['E'].width = 16

        wb.save(file_path)
        
        # Set task download filename
        tasks[task_id]['download_name'] = custom_excel_filename
        log_task(task_id, f"Successfully created segmented professional Excel workbook: {custom_excel_filename}", "success")        # Sync to Google Sheets if configured
        log_task(task_id, f"Checking Google Sheet URL configuration: '{google_sheet_url}'", "info")
        if google_sheet_url:
            try:
                log_task(task_id, "Syncing dataset to Google Sheet...", "info")
                import urllib.request
                import json
                
                records = []
                for _, row in df_export.iterrows():
                    records.append({
                        "date": str(row['Date']),
                        "close_0928": float(row['09:28 Close']),
                        "callStrike": int(row['Call Strike']),
                        "putStrike": int(row['Put Strike']),
                        # 15m Heikin Ashi values
                        "callPrev1515Close15mHA": float(row['Call Prev 15:15 Close (15m HA)']) if row['Call Prev 15:15 Close (15m HA)'] != "" else "",
                        "call0915Close15mHA": float(row['Call 09:15 Close (15m HA)']) if row['Call 09:15 Close (15m HA)'] != "" else "",
                        "call0930Close15mHA": float(row['Call 09:30 Close (15m HA)']) if row['Call 09:30 Close (15m HA)'] != "" else "",
                        "call0945Open15mHA": float(row['Call 09:45 Open (15m HA)']) if row['Call 09:45 Open (15m HA)'] != "" else "",
                        "putPrev1515Close15mHA": float(row['Put Prev 15:15 Close (15m HA)']) if row['Put Prev 15:15 Close (15m HA)'] != "" else "",
                        "put0915Close15mHA": float(row['Put 09:15 Close (15m HA)']) if row['Put 09:15 Close (15m HA)'] != "" else "",
                        "put0930Close15mHA": float(row['Put 09:30 Close (15m HA)']) if row['Put 09:30 Close (15m HA)'] != "" else "",
                        "put0945Open15mHA": float(row['Put 09:45 Open (15m HA)']) if row['Put 09:45 Open (15m HA)'] != "" else ""
                    })
                
                payload = json.dumps(records).encode('utf-8')
                req = urllib.request.Request(
                    google_sheet_url,
                    data=payload,
                    headers={
                        'Content-Type': 'application/json',
                        'User-Agent': 'Mozilla/5.0'
                    },
                    method='POST'
                )
                with urllib.request.urlopen(req, timeout=25) as res:
                    res_data = json.loads(res.read().decode('utf-8'))
                    if res_data.get('status') == 'success':
                        log_task(task_id, f"Google Sheet synced successfully! Added {res_data.get('rows_added', len(records))} rows.", "success")
                    else:
                        log_task(task_id, f"Google Sheet sync error: {res_data.get('error', 'unknown')}", "error")
            except Exception as e:
                log_task(task_id, f"Failed to sync to Google Sheet: {str(e)}", "error")
        
        # Stream a clean table preview of the trailing data rows
        preview_text = "PREVIEW:\n" + df_export.tail(5).to_string(index=False)
        log_task(task_id, preview_text, "info")
        
        with tasks_lock:
            tasks[task_id]['status'] = 'completed'
            tasks[task_id]['file_path'] = file_path
            tasks[task_id]['download_name'] = filename
            
    except Exception as e:
        error_msg = str(e)
        log_task(task_id, f"Error: {error_msg}", "error")
        with tasks_lock:
            tasks[task_id]['status'] = 'failed'
            tasks[task_id]['error'] = error_msg

# Web Server Routes
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/extract', methods=['POST'])
def start_extraction():
    data = request.json or {}
    
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    tv_session = data.get('tv_session', '').strip()  # Pre-authenticated session token
    symbol = data.get('symbol', '').strip().upper()
    exchange = data.get('exchange', '').strip().upper()
    interval = data.get('interval', '1 Minute')
    n_bars = data.get('n_bars', 1000)
    filename = data.get('filename', 'export.xlsx').strip()
    time_filter = data.get('time_filter', 'all').strip()
    strike_offset = data.get('strike_offset', 100)
    google_sheet_url = data.get('google_sheet_url', '').strip()
    live_today_only = bool(data.get('live_today_only', False))
    
    baseline_interval = data.get('baseline_interval', '15 Minutes').strip()
    signal_interval = data.get('signal_interval', '5 Minutes').strip()
    
    if not symbol or not exchange:
        return jsonify({"error": "Symbol and Exchange are required"}), 400
        
    try:
        n_bars = int(n_bars)
        if n_bars <= 0:
            raise ValueError()
    except ValueError:
        return jsonify({"error": "Number of bars must be a positive integer"}), 400
    
    try:
        strike_offset = int(strike_offset)
        if strike_offset <= 0:
            raise ValueError()
    except (ValueError, TypeError):
        strike_offset = 100
        
    if not filename:
        filename = f"{symbol}_data.xlsx"

    task_id = str(uuid.uuid4())
    
    # Store task details
    with tasks_lock:
        tasks[task_id] = {
            'status': 'running',
            'logs': [],
            'file_path': None,
            'download_name': None,
            'error': None,
            'preview_summary': [],
            'preview_live_5m': []
        }
        
    # Start extraction thread
    thread = threading.Thread(
        target=extraction_thread,
        args=(task_id, username, password, symbol, exchange, interval, n_bars, filename, time_filter, strike_offset, google_sheet_url, live_today_only, tv_session),
        kwargs={'baseline_interval_name': baseline_interval, 'signal_interval_name': signal_interval},
        daemon=True
    )
    thread.start()
    
    return jsonify({"task_id": task_id})


@app.route('/api/task/<task_id>/symbols')
def get_task_symbols(task_id):
    """Return the dynamically computed call/put symbols for a running or completed task."""
    with tasks_lock:
        if task_id not in tasks:
            return jsonify({"error": "Task not found"}), 404
        return jsonify({
            "status":  tasks[task_id]['status'],
            "symbols": tasks[task_id].get('symbols', [])
        })


@app.route('/api/preview_symbols', methods=['POST'])
def preview_symbols():
    """
    Fast endpoint: fetches only the NIFTY underlying 1-min data,
    computes the call/put option symbols for each date from the 09:28 bar,
    and returns them WITHOUT fetching any option contracts.
    Used by the UI to show the symbol preview panel before/during extraction.
    """
    data = request.json or {}
    username      = data.get('username', '').strip()
    password      = data.get('password', '').strip()
    symbol        = data.get('symbol', '').strip().upper()
    exchange      = data.get('exchange', '').strip().upper()
    n_bars        = max(int(data.get('n_bars', 5000)), 5000)
    strike_offset = int(data.get('strike_offset', 100))

    if not symbol or not exchange:
        return jsonify({"error": "Symbol and Exchange are required"}), 400

    try:
        if username and password:
            tv = TvDatafeed(username=username, password=password)
        else:
            tv = TvDatafeed()

        raw = safe_get_hist(tv=tv, symbol=symbol, exchange=exchange,
                            interval=Interval.in_1_minute, n_bars=n_bars)
        if raw is None or raw.empty:
            return jsonify({"error": "No data returned for symbol"}), 400

        raw = raw.sort_index()
        raw.index.name = 'datetime'
        df = raw.reset_index()
        df['datetime'] = pd.to_datetime(df['datetime'])
        df['ist_date'] = df['datetime'].dt.strftime('%Y-%m-%d')
        df['ist_time'] = df['datetime'].dt.strftime('%H:%M')

        MONTH_CODE = {1:'1',2:'2',3:'3',4:'4',5:'5',6:'6',
                      7:'7',8:'8',9:'9',10:'O',11:'N',12:'D'}

        # Target day calculation matching extraction logic
        def weekly_expiry(date_str):
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            sym_upper = symbol.upper()
            if "FINNIFTY" in sym_upper:
                target_weekday = 1 # Tuesday
            elif "BANKNIFTY" in sym_upper or "NIFTYBANK" in sym_upper:
                target_weekday = 2 # Wednesday
            else:
                target_weekday = 3 # Thursday for NIFTY
                
            days_ahead = target_weekday - dt.weekday()
            if days_ahead < 0:
                days_ahead += 7
            exp = dt + timedelta(days=days_ahead)
            
            return exp.strftime("%y%m%d"), exp.strftime("%d %b %Y")

        rows = []
        for current_date in sorted(df['ist_date'].unique()):
            bar = df[(df['ist_date'] == current_date) & (df['ist_time'] == '09:28')]
            if bar.empty:
                continue
            close = float(bar['close'].iloc[0])
            call_strike = int(((close - strike_offset) // strike_offset) * strike_offset)
            put_strike  = int(np.ceil((close + strike_offset) / float(strike_offset)) * strike_offset)
            exp_str, exp_label = weekly_expiry(current_date)
            rows.append({
                'date':         current_date,
                'close_0928':   round(close, 2),
                'expiry':       exp_str,
                'expiry_label': exp_label,
                'call_strike':  call_strike,
                'put_strike':   put_strike,
                'call_sym':     f"{symbol}{exp_str}{call_strike}CE",
                'put_sym':      f"{symbol}{exp_str}{put_strike}PE"
            })

        return jsonify({"symbols": rows})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/task/<task_id>')
def get_task_status(task_id):
    since_idx = int(request.args.get('since', 0))
    
    with tasks_lock:
        if task_id not in tasks:
            return jsonify({"error": "Task not found"}), 404
            
        task = tasks[task_id]
        all_logs = task['logs']
        new_logs = all_logs[since_idx:] if since_idx < len(all_logs) else []
        
        response_data = {
            "status": task['status'],
            "new_logs": new_logs,
            "error": task['error'],
            "preview_summary": task.get('preview_summary', []),
            "preview_live_5m": task.get('preview_live_5m', [])
        }
        
        return jsonify(response_data)

@app.route('/api/download/<task_id>')
def download_file(task_id):
    with tasks_lock:
        if task_id not in tasks:
            return "Task not found", 404
            
        task = tasks[task_id]
        if task['status'] != 'completed' or not task['file_path']:
            return "File not ready or job failed", 400
            
        file_path = task['file_path']
        download_name = task['download_name']
        
    return send_file(file_path, as_attachment=True, download_name=download_name)

@app.route('/analyzer')
def analyzer_home():
    return render_template('analyzer.html')

@app.route('/api/tasks', methods=['GET'])
def list_tasks():
    rows = []
    with tasks_lock:
        for tid, tdata in tasks.items():
            symbol_info = tdata.get('symbols', [])
            latest_sym = symbol_info[-1] if symbol_info else None
            
            rows.append({
                "task_id": tid,
                "status": tdata.get('status', ''),
                "date": latest_sym.get('date', '') if latest_sym else '',
                "call_sym": latest_sym.get('call_sym', '') if latest_sym else '',
                "put_sym": latest_sym.get('put_sym', '') if latest_sym else '',
                "close_0928": latest_sym.get('close_0928', '') if latest_sym else ''
            })
            
    return jsonify({"tasks": rows})

# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
# Shared Backtest Engine Helper
# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
def _run_backtest_engine(candles, baseline, target_points, sl_type):
    """Core single-side backtest engine. Returns full result dict with critic metrics."""
    # Slicing Rule: Completely ignore the 9:15 AM and 9:20 AM candles
    filtered_candles = []
    for c in candles:
        time_str = str(c['time']).strip()
        if ' ' in time_str:
            time_str = time_str.split(' ')[-1]
        if len(time_str.split(':')) > 2:
            time_str = ':'.join(time_str.split(':')[:2])
        if time_str in ['09:15', '09:20']:
            continue
        c_copy = c.copy()
        c_copy['time'] = time_str
        filtered_candles.append(c_copy)
    candles = filtered_candles

    # 1. Setup Candle Scanner: Monitor candles for a candle completely below baseline (High < Baseline)
    setup_candle_idx = None
    locked_sl = None
    for i, c in enumerate(candles):
        if c['high'] < baseline:
            setup_candle_idx = i
            locked_sl = c['low']
            break

    # 2. Entry Breach Scanner: Scan for 2nd breach of baseline
    first_breach_idx = None
    entry_idx = None
    entry_price = None
    entry_time = None
    entry_type = "auto"

    for i, c in enumerate(candles):
        high_val = max(c['open'], c['high'], c['low'], c['close'])
        if high_val > baseline:
            if first_breach_idx is None:
                first_breach_idx = i
            elif entry_idx is None:
                entry_idx = i
                entry_price = c['open']
                entry_time = c['time']
                break

    # Fallback: only first breach, no confirmation candle
    if first_breach_idx is not None and entry_idx is None:
        c = candles[first_breach_idx]
        entry_idx = first_breach_idx
        entry_price = max(c['open'], c['high'], c['low'], c['close'])
        entry_time = c['time']
        entry_type = "fallback"

    target_val = None
    target_hit_idx = None
    target_hit_time = None
    target_exit_price = None
    sl_hit_idx = None
    sl_hit_time = None
    sl_exit_price = None

    if entry_idx is not None:
        entry_ref_high = candles[entry_idx]['high']
        target_val = entry_ref_high + target_points
        for i in range(entry_idx, len(candles)):
            c = candles[i]
            if c['high'] >= target_val:
                target_hit_idx = i
                target_hit_time = c['time']
                target_exit_price = target_val
                break
        
        # Stop Loss Exit Scanner: Active ONLY if setup candle has been identified
        if setup_candle_idx is not None:
            start_scan_idx = max(entry_idx, setup_candle_idx + 1)
            for i in range(start_scan_idx, len(candles)):
                c = candles[i]
                trigger_val = c['low'] if sl_type == 'low' else c['close']
                if trigger_val < locked_sl:
                    sl_hit_idx = i
                    sl_hit_time = c['time']
                    sl_exit_price = min(c['open'], c['high'], c['low'], c['close'])
                    break

    exit_reason = "Open/No Exit"
    exit_time = None
    exit_price = None
    pnl = 0.0

    if entry_idx is not None:
        if target_hit_idx is not None and sl_hit_idx is not None:
            if target_hit_idx <= sl_hit_idx:
                exit_reason, exit_time, exit_price = "Target Hit", target_hit_time, target_exit_price
            else:
                exit_reason, exit_time, exit_price = "Stop Loss Hit", sl_hit_time, sl_exit_price
        elif target_hit_idx is not None:
            exit_reason, exit_time, exit_price = "Target Hit", target_hit_time, target_exit_price
        elif sl_hit_idx is not None:
            exit_reason, exit_time, exit_price = "Stop Loss Hit", sl_hit_time, sl_exit_price
        if exit_price is not None:
            pnl = round(exit_price - entry_price, 2)

    # Ã¢â€â‚¬Ã¢â€â‚¬ Critic Metrics Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    critic = {}
    if entry_idx is not None and entry_price is not None:
        entry_above_pct = round((entry_price - baseline) / baseline * 100, 2) if baseline > 0 else 0
        sl_distance = entry_price - baseline
        target_distance = (target_val - entry_price) if target_val else 0
        rr = round(target_distance / sl_distance, 2) if sl_distance > 0 else None

        post_entry = candles[entry_idx:]
        mfe_price = max((c['high'] for c in post_entry), default=entry_price)
        mae_price = min((c['low'] for c in post_entry), default=entry_price)
        mfe_pts = round(mfe_price - entry_price, 2)
        mae_pts = round(entry_price - mae_price, 2)

        exit_idx = target_hit_idx or sl_hit_idx
        candles_in_trade = (exit_idx - entry_idx + 1) if exit_idx else len(candles) - entry_idx
        efficiency = round(pnl / mfe_pts * 100, 1) if mfe_pts > 0 and pnl > 0 else 0.0

        # Composite score (0-100)
        score = 0
        if entry_above_pct <= 3:  score += 25
        elif entry_above_pct <= 8:  score += 18
        elif entry_above_pct <= 15: score += 10
        else:                       score += 2

        if rr is not None:
            if rr >= 2.0:   score += 25
            elif rr >= 1.5: score += 18
            elif rr >= 1.0: score += 10
            else:           score += 3

        if entry_idx <= 1:   score += 20
        elif entry_idx <= 3: score += 14
        elif entry_idx <= 6: score += 8
        else:                score += 2

        if efficiency >= 80: score += 15
        elif efficiency >= 55: score += 10
        elif efficiency >= 30: score += 5

        if mae_pts < entry_price * 0.03:   score += 15
        elif mae_pts < entry_price * 0.07: score += 8
        else:                               score += 2

        grade = 'A' if score >= 82 else 'B' if score >= 66 else 'C' if score >= 50 else 'D' if score >= 35 else 'F'

        critic = {
            'entry_above_pct': entry_above_pct,
            'signal_speed_candles': entry_idx,
            'first_breach_idx': first_breach_idx,
            'risk_reward': rr,
            'mfe': round(mfe_price, 2),
            'mfe_points': mfe_pts,
            'mae': round(mae_price, 2),
            'mae_points': mae_pts,
            'efficiency_pct': efficiency,
            'candles_in_trade': candles_in_trade,
            'score': score,
            'grade': grade,
        }

    # Tag candles for UI highlighting
    for i, c in enumerate(candles):
        c['tag'] = ''
        if first_breach_idx is not None and i == first_breach_idx and i != entry_idx:
            c['tag'] = 'breach'
        if entry_idx is not None and i == entry_idx:
            c['tag'] = 'entry'
        if exit_reason == 'Target Hit' and target_hit_idx is not None and i == target_hit_idx:
            c['tag'] = 'target'
        elif exit_reason == 'Stop Loss Hit' and sl_hit_idx is not None and i == sl_hit_idx:
            c['tag'] = 'stop_loss'

    return {
        'baseline': baseline,
        'first_breach_idx': first_breach_idx,
        'entry_idx': entry_idx,
        'entry_time': entry_time,
        'entry_price': entry_price,
        'entry_type': entry_type,
        'target_value': target_val,
        'target_hit_time': target_hit_time,
        'sl_value': locked_sl if locked_sl is not None else baseline,
        'sl_hit_time': sl_hit_time,
        'sl_exit_price': sl_exit_price,
        'final_exit_reason': exit_reason,
        'final_exit_time': exit_time,
        'final_exit_price': exit_price,
        'pnl': pnl,
        'candles': candles,
        'critic': critic,
    }


def _generate_critic_commentary(critic, winner_side, opp_first_breach_idx, opp_entry_idx):
    """Generate qualitative analysis text for the winning trade."""
    lines = []
    entry_pct  = critic.get('entry_above_pct', 0)
    rr         = critic.get('risk_reward')
    speed      = critic.get('signal_speed_candles', 0)
    eff        = critic.get('efficiency_pct', 0)
    mae_pts    = critic.get('mae_points', 0)
    mfe_pts    = critic.get('mfe_points', 0)
    entry_type = critic.get('entry_type', 'auto')

    # Entry quality
    if entry_pct <= 3:
        lines.append(f"Entry price was only {entry_pct:.1f}% above baseline Ã¢â‚¬â€ a textbook tight entry with minimal premium paid over the reference level.")
    elif entry_pct <= 8:
        lines.append(f"Entry price was {entry_pct:.1f}% above baseline Ã¢â‚¬â€ clean entry, close to the reference level.")
    elif entry_pct <= 15:
        lines.append(f"Entry price was {entry_pct:.1f}% above baseline Ã¢â‚¬â€ acceptable, though some premium was paid above the reference.")
    else:
        lines.append(f"Entry price was {entry_pct:.1f}% above baseline Ã¢â‚¬â€ the market had already stretched considerably before entry, elevating risk.")

    # Entry type note
    if entry_type == 'fallback':
        lines.append("Entry was triggered via the fallback branch (only one breach candle found, no confirmation) Ã¢â‚¬â€ treat with slightly lower confidence.")

    # Signal speed
    candle_labels = ['the very first candle', 'the 2nd candle', 'the 3rd candle', 'the 4th candle', 'the 5th candle', 'the 6th candle']
    label = candle_labels[min(speed, len(candle_labels)-1)] if speed < len(candle_labels) else f"candle #{speed+1}"
    if speed <= 1:
        lines.append(f"Signal confirmed at {label} from market open Ã¢â‚¬â€ exceptional early momentum.")
    elif speed <= 3:
        lines.append(f"Signal confirmed at {label} from market open Ã¢â‚¬â€ strong directional conviction.")
    elif speed <= 6:
        lines.append(f"Signal confirmed at {label} from market open Ã¢â‚¬â€ normal intraday development.")
    else:
        lines.append(f"Signal confirmed late at {label} from market open Ã¢â‚¬â€ delayed momentum suggests weaker trend conviction.")

    # Risk-reward
    if rr is not None:
        if rr >= 2.0:
            lines.append(f"Risk-reward of {rr:.2f}:1 is excellent Ã¢â‚¬â€ this is a high-quality setup with clear positive expected value.")
        elif rr >= 1.5:
            lines.append(f"Risk-reward of {rr:.2f}:1 is good Ã¢â‚¬â€ the setup had a healthy reward relative to its risk.")
        elif rr >= 1.0:
            lines.append(f"Risk-reward of {rr:.2f}:1 is borderline Ã¢â‚¬â€ reward barely covers risk; tighter entries would improve this.")
        else:
            lines.append(f"Risk-reward of {rr:.2f}:1 is poor Ã¢â‚¬â€ the stop-loss distance was too wide relative to target, reducing trade quality.")

    # Directional dominance
    if opp_first_breach_idx is None:
        lines.append(f"The opposing side showed zero breach throughout the entire session Ã¢â‚¬â€ outstanding directional conviction for the {winner_side.upper()} trade.")
    elif opp_entry_idx is None:
        lines.append(f"The opposing side showed a first breach but never confirmed entry Ã¢â‚¬â€ the {winner_side.upper()} had clear directional dominance.")
    else:
        gap = (opp_entry_idx or 99) - speed
        if gap >= 6:
            lines.append(f"The {winner_side.upper()} side entered {gap} candles ahead of any opposing confirmation Ã¢â‚¬â€ strong edge and directional conviction.")
        elif gap >= 2:
            lines.append(f"The {winner_side.upper()} side led by {gap} candles before the opposing side stirred Ã¢â‚¬â€ moderate directional confidence.")
        else:
            lines.append("Both sides were racing closely Ã¢â‚¬â€ this was a tightly contested session; manage risk carefully on such days.")

    # Trade efficiency
    if mfe_pts > 0:
        if eff >= 80:
            lines.append(f"Trade captured {eff:.1f}% of its maximum favorable move ({mfe_pts:.1f} pts available) Ã¢â‚¬â€ highly efficient exit.")
        elif eff >= 55:
            lines.append(f"Trade captured {eff:.1f}% of its maximum favorable move ({mfe_pts:.1f} pts available) Ã¢â‚¬â€ solid exit efficiency.")
        elif eff > 0:
            lines.append(f"Trade captured only {eff:.1f}% of its maximum favorable move ({mfe_pts:.1f} pts available) Ã¢â‚¬â€ significant upside remained unexploited.")
        else:
            lines.append(f"Trade did not reach target; maximum upside was {mfe_pts:.1f} pts above entry.")

    # Drawdown
    if mae_pts > 0:
        if mae_pts < 5:
            lines.append(f"Maximum adverse excursion was only {mae_pts:.1f} pts Ã¢â‚¬â€ minimal drawdown, price barely tested the entry.")
        elif mae_pts < 15:
            lines.append(f"Maximum adverse excursion was {mae_pts:.1f} pts Ã¢â‚¬â€ manageable drawdown, stop-loss was not threatened critically.")
        else:
            lines.append(f"Maximum adverse excursion reached {mae_pts:.1f} pts Ã¢â‚¬â€ significant drawdown; position sizing discipline is essential on such trades.")

    return " ".join(lines)


@app.route('/api/analyze_dual', methods=['POST'])
def analyze_dual():
    """Parallel CALL + PUT First-Signal-Wins strategy engine with critic analysis."""
    try:
        target_points = float(request.form.get('target_points', 25))
        sl_type = request.form.get('sl_type', 'close').lower()
        task_id = request.form.get('task_id', '').strip() or None
        file = request.files.get('file')

        call_candles, put_candles = [], []
        call_bl_info, put_bl_info = {}, {}

        if task_id:
            with tasks_lock:
                if task_id not in tasks:
                    return jsonify({"error": f"Task '{task_id}' not found."}), 404
                task_data = tasks[task_id]

            preview_summary = task_data.get('preview_summary', [])
            preview_live_5m = task_data.get('preview_live_5m', [])
            if not preview_summary:
                return jsonify({"error": "Task has no strategy summary data yet."}), 400

            last_row = preview_summary[-1]
            c_prev  = float(last_row.get('Call Prev 15:15 Close (15m HA)', 0) or 0)
            c_0915  = float(last_row.get('Call 09:15 Close (15m HA)', 0) or 0)
            p_prev  = float(last_row.get('Put Prev 15:15 Close (15m HA)', 0) or 0)
            p_0915  = float(last_row.get('Put 09:15 Close (15m HA)', 0) or 0)
            call_bl = max(c_prev, c_0915)
            put_bl  = max(p_prev, p_0915)
            call_bl_info = {'prev_day': c_prev, 'today_0915': c_0915, 'baseline': call_bl}
            put_bl_info  = {'prev_day': p_prev, 'today_0915': p_0915, 'baseline': put_bl}

            for row in preview_live_5m:
                ts = row.get('Time (IST)', '')
                if not ts: continue
                try:
                    if row.get('Call Open', '') != '':
                        call_candles.append({'time': ts, 'open': float(row['Call Open']),
                            'high': float(row['Call High']), 'low': float(row['Call Low']), 'close': float(row['Call Close'])})
                    if row.get('Put Open', '') != '':
                        put_candles.append({'time': ts, 'open': float(row['Put Open']),
                            'high': float(row['Put High']), 'low': float(row['Put Low']), 'close': float(row['Put Close'])})
                except (ValueError, KeyError):
                    continue

        elif file and file.filename != '':
            import pandas as pd
            xls = pd.ExcelFile(file)
            if 'Strategy_Summary' not in xls.sheet_names or '5m_Live_Candles' not in xls.sheet_names:
                return jsonify({"error": "File must contain 'Strategy_Summary' and '5m_Live_Candles' sheets."}), 400
            df_s = pd.read_excel(xls, 'Strategy_Summary')
            df_c = pd.read_excel(xls, '5m_Live_Candles')
            lr = df_s.iloc[-1]
            c_prev = float(lr.get('Call Prev 15:15 Close (15m HA)', 0) or 0)
            c_0915 = float(lr.get('Call 09:15 Close (15m HA)', 0) or 0)
            p_prev = float(lr.get('Put Prev 15:15 Close (15m HA)', 0) or 0)
            p_0915 = float(lr.get('Put 09:15 Close (15m HA)', 0) or 0)
            call_bl = max(c_prev, c_0915)
            put_bl  = max(p_prev, p_0915)
            call_bl_info = {'prev_day': c_prev, 'today_0915': c_0915, 'baseline': call_bl}
            put_bl_info  = {'prev_day': p_prev, 'today_0915': p_0915, 'baseline': put_bl}
            tc = next((c for c in df_c.columns if 'Time' in c), df_c.columns[0])
            for _, row in df_c.iterrows():
                if pd.isna(row[tc]): continue
                rt = row[tc]
                ts = rt.strftime('%H:%M') if hasattr(rt, 'strftime') else ':'.join(str(rt).strip().split(':')[:2])
                try:
                    if not pd.isna(row.get('Call Open', float('nan'))):
                        call_candles.append({'time': ts, 'open': float(row['Call Open']),
                            'high': float(row['Call High']), 'low': float(row['Call Low']), 'close': float(row['Call Close'])})
                    if not pd.isna(row.get('Put Open', float('nan'))):
                        put_candles.append({'time': ts, 'open': float(row['Put Open']),
                            'high': float(row['Put High']), 'low': float(row['Put Low']), 'close': float(row['Put Close'])})
                except (ValueError, KeyError):
                    continue
        else:
            return jsonify({"error": "Provide a task_id or upload an Excel file."}), 400

        if not call_candles and not put_candles:
            return jsonify({"error": "No candle data found for either CALL or PUT."}), 400

        # Ã¢â€â‚¬Ã¢â€â‚¬ Shared-Timeline Race Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
        c_first = None   # CALL first breach index
        p_first = None   # PUT  first breach index
        c_entry = None   # CALL entry index
        p_entry = None   # PUT  entry index
        winner  = 'none'

        for i in range(max(len(call_candles), len(put_candles))):
            if c_entry is None and i < len(call_candles):
                cc = call_candles[i]
                if max(cc['open'], cc['high'], cc['low'], cc['close']) > call_bl:
                    if c_first is None: c_first = i
                    else: c_entry = i

            if p_entry is None and i < len(put_candles):
                pc = put_candles[i]
                if max(pc['open'], pc['high'], pc['low'], pc['close']) > put_bl:
                    if p_first is None: p_first = i
                    else: p_entry = i

            if c_entry is not None and p_entry is None:   winner = 'call'; break
            if p_entry is not None and c_entry is None:   winner = 'put';  break
            if c_entry is not None and p_entry is not None:
                winner = 'call' if c_entry <= p_entry else 'put'; break

        # Ã¢â€â‚¬Ã¢â€â‚¬ Full backtest for BOTH sides (winner gets full critic) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
        call_res = _run_backtest_engine([c.copy() for c in call_candles], call_bl, target_points, sl_type)
        put_res  = _run_backtest_engine([c.copy() for c in put_candles],  put_bl,  target_points, sl_type)

        # Determine status labels
        call_status = 'entered' if c_entry is not None else ('cancelled' if winner == 'put' else 'no_entry')
        put_status  = 'entered' if p_entry is not None else ('cancelled' if winner == 'call' else 'no_entry')
        call_res['status'] = call_status
        put_res['status']  = put_status

        # If cancelled, override PnL to zero and mark candles correctly
        if call_status == 'cancelled':
            call_res['pnl'] = 0.0
            call_res['final_exit_reason'] = 'Cancelled Ã¢â‚¬â€ PUT Entered First'
            call_res['critic'] = {}
        if put_status == 'cancelled':
            put_res['pnl'] = 0.0
            put_res['final_exit_reason'] = 'Cancelled Ã¢â‚¬â€ CALL Entered First'
            put_res['critic'] = {}

        # Augment winner critic with cross-side directional dominance
        trade_res = call_res if winner == 'call' else put_res if winner == 'put' else None
        if trade_res and trade_res.get('critic'):
            opp_first = p_first if winner == 'call' else c_first
            opp_entry = p_entry if winner == 'call' else c_entry
            trade_res['critic']['opp_first_breach_idx'] = opp_first
            trade_res['critic']['opp_entry_idx']        = opp_entry
            trade_res['critic']['dominance_gap']        = (
                (opp_entry - (c_entry if winner == 'call' else p_entry))
                if opp_entry is not None and (c_entry if winner == 'call' else p_entry) is not None else None
            )
            # Generate commentary
            trade_res['critic']['commentary'] = _generate_critic_commentary(
                trade_res['critic'], winner, opp_first, opp_entry
            )

        def _side_dict(res, bl_info, status):
            return {
                'status': status,
                'baseline': bl_info['baseline'],
                'prev_day': bl_info['prev_day'],
                'today_0915': bl_info['today_0915'],
                'entry_time': res.get('entry_time'),
                'entry_price': res.get('entry_price'),
                'entry_type': res.get('entry_type'),
                'first_breach_idx': res.get('first_breach_idx'),
                'final_exit_reason': res.get('final_exit_reason'),
                'pnl': res.get('pnl', 0.0),
                'candles': res.get('candles', []),
                'critic': res.get('critic', {}),
            }

        # Resolve the winning contract symbol from task database
        winner_sym = None
        if task_id:
            with tasks_lock:
                task_data = tasks.get(task_id)
                if task_data:
                    symbols = task_data.get('symbols', [])
                    latest_sym = symbols[-1] if symbols else None
                    if latest_sym:
                        winner_sym = latest_sym.get('call_sym') if winner == 'call' else latest_sym.get('put_sym')

        # Check settings for auto-pilot
        import os, json
        settings_path = os.path.join(os.path.dirname(__file__), 'settings.json')
        auto_pilot = False
        lot_size = 50
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r', encoding='utf-8') as sf:
                    sdata = json.load(sf)
                    lot_size = int(sdata.get('lot_size', 50))
                    if sdata.get('auto_pilot_enabled') and sdata.get('safety_lock_key') == "I UNDERSTAND THE RISK":
                        auto_pilot = True
            except:
                pass

        # Build signal events for the winning trade
        winner_res = call_res if winner == 'call' else put_res if winner == 'put' else None
        winner_bl  = call_bl if winner == 'call' else put_bl if winner == 'put' else 0
        trade_signal_events = []
        if winner_res and winner_res.get('entry_idx') is not None:
            trade_signal_events = _build_signal_events(
                winner_res.get('candles', []),
                winner_bl,
                winner_res.get('entry_idx'),
                winner_res.get('entry_time'),
                winner_res.get('entry_price'),
                winner_res.get('entry_type', 'auto'),
                winner_res.get('first_breach_idx'),
                winner_res.get('final_exit_reason', 'Open/No Exit'),
                winner_res.get('final_exit_time'),
                winner_res.get('final_exit_price'),
                winner_res.get('target_value'),
                winner_res.get('sl_value', winner_bl),
                winner,
                lot_size=lot_size,
                tradingsymbol=winner_sym
            )

        # Trigger auto-pilot execution if enabled
        if auto_pilot and winner_res and winner_res.get('entry_idx') is not None:
            _run_autopilot_execution(trade_signal_events, winner_sym, trade_res.get('pnl', 0.0) if trade_res else 0.0)

            # Update event status to show they are executed via Auto-Pilot
            for ev in trade_signal_events:
                if ev.get('actionable') and ev.get('time') != 'Ã¢â‚¬â€':
                    ev['status'] = 'Ã¢Å¡Â¡ AUTO-EXECUTED'

        return jsonify({
            'winner': winner,
            'call': _side_dict(call_res, call_bl_info, call_status),
            'put':  _side_dict(put_res,  put_bl_info,  put_status),
            'trade': {
                'option_type':       winner,
                'entry_time':        trade_res.get('entry_time')        if trade_res else None,
                'entry_price':       trade_res.get('entry_price')       if trade_res else None,
                'entry_type':        trade_res.get('entry_type')        if trade_res else None,
                'target_value':      trade_res.get('target_value')      if trade_res else None,
                'target_hit_time':   trade_res.get('target_hit_time')   if trade_res else None,
                'sl_value':          trade_res.get('sl_value')          if trade_res else None,
                'sl_hit_time':       trade_res.get('sl_hit_time')       if trade_res else None,
                'final_exit_reason': trade_res.get('final_exit_reason') if trade_res else 'No Entry Today',
                'final_exit_time':   trade_res.get('final_exit_time')   if trade_res else None,
                'final_exit_price':  trade_res.get('final_exit_price')  if trade_res else None,
                'pnl':               trade_res.get('pnl', 0.0)          if trade_res else 0.0,
                'critic':            trade_res.get('critic', {})         if trade_res else {},
                'signal_events':     trade_signal_events,
            }
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/analyze', methods=['POST'])
def analyze_strategy():
    try:
        # Determine if we are processing an uploaded file or manual data
        candles = []
        baseline_vals = {}
        
        target_points = float(request.form.get('target_points', 25))
        sl_type = request.form.get('sl_type', 'close').lower() # 'close' or 'low'
        option_type = request.form.get('option_type', 'call').lower() # 'call' or 'put'
        
        # Manual overrides (empty strings/null are ignored)
        def get_float_or_none(key):
            val = request.form.get(key, '')
            if val.strip() == '':
                return None
            try:
                return float(val)
            except ValueError:
                return None

        manual_baseline = get_float_or_none('manual_baseline')
        manual_entry_price = get_float_or_none('manual_entry_price')
        manual_entry_time = request.form.get('manual_entry_time', '').strip() or None
        manual_target = get_float_or_none('manual_target')
        manual_sl = get_float_or_none('manual_sl')
        manual_exit_time = request.form.get('manual_exit_time', '').strip() or None

        task_id = request.form.get('task_id', '').strip() or None

        # Check if file uploaded or task_id provided
        file = request.files.get('file')
        
        if task_id:
            # Load directly from task state in memory
            with tasks_lock:
                if task_id not in tasks:
                    return jsonify({"error": f"Task ID '{task_id}' not found in active jobs list."}), 404
                task_data = tasks[task_id]
                
            preview_summary = task_data.get('preview_summary', [])
            preview_live_5m = task_data.get('preview_live_5m', [])
            
            if not preview_summary:
                return jsonify({"error": f"The selected task '{task_id}' has not processed any Strategy Summary rows yet."}), 400
                
            # Get latest summary row
            last_row = preview_summary[-1]
            
            # Find the required baseline columns
            if option_type == 'call':
                prev_day_key = 'Call Prev 15:15 Close (15m HA)'
                today_0915_key = 'Call 09:15 Close (15m HA)'
            else:
                prev_day_key = 'Put Prev 15:15 Close (15m HA)'
                today_0915_key = 'Put 09:15 Close (15m HA)'
                
            if prev_day_key not in last_row or today_0915_key not in last_row:
                return jsonify({"error": f"Required 15m HA keys not found in selected task data."}), 400
                
            baseline_vals['prev_day_1515_close_ha'] = float(last_row[prev_day_key]) if last_row[prev_day_key] != "" else 0.0
            baseline_vals['today_0915_close_ha'] = float(last_row[today_0915_key]) if last_row[today_0915_key] != "" else 0.0
            
            # Extract 5m standard candles
            open_key = 'Call Open' if option_type == 'call' else 'Put Open'
            high_key = 'Call High' if option_type == 'call' else 'Put High'
            low_key = 'Call Low' if option_type == 'call' else 'Put Low'
            close_key = 'Call Close' if option_type == 'call' else 'Put Close'
            
            for row in preview_live_5m:
                time_str = row.get('Time (IST)', '')
                if not time_str or open_key not in row or row[open_key] == "":
                    continue
                candles.append({
                    "time": time_str,
                    "open": float(row[open_key]),
                    "high": float(row[high_key]),
                    "low": float(row[low_key]),
                    "close": float(row[close_key])
                })
        elif file and file.filename != '':
            # Load using pandas
            import pandas as pd
            xls = pd.ExcelFile(file)
            
            if 'Strategy_Summary' not in xls.sheet_names or '5m_Live_Candles' not in xls.sheet_names:
                return jsonify({"error": "Uploaded file must contain 'Strategy_Summary' and '5m_Live_Candles' sheets."}), 400
                
            df_summary = pd.read_excel(xls, 'Strategy_Summary')
            df_candles = pd.read_excel(xls, '5m_Live_Candles')
            
            # 1. Parse Baseline reference from Strategy_Summary (last row)
            if df_summary.empty:
                return jsonify({"error": "Strategy_Summary sheet is empty."}), 400
                
            last_row = df_summary.iloc[-1]
            
            # Find the required baseline columns
            if option_type == 'call':
                prev_day_col = 'Call Prev 15:15 Close (15m HA)'
                today_0915_col = 'Call 09:15 Close (15m HA)'
            else:
                prev_day_col = 'Put Prev 15:15 Close (15m HA)'
                today_0915_col = 'Put 09:15 Close (15m HA)'
                
            if prev_day_col not in df_summary.columns or today_0915_col not in df_summary.columns:
                # Fallback to column prefix matching in case names differ slightly
                prev_day_col = next((c for c in df_summary.columns if 'Prev 15:15 Close' in c and option_type.upper() in c.upper()), None)
                today_0915_col = next((c for c in df_summary.columns if '09:15 Close' in c and option_type.upper() in c.upper()), None)
                
            if not prev_day_col or not today_0915_col:
                return jsonify({"error": f"Could not find required 15m HA columns for {option_type.upper()} in Strategy_Summary sheet."}), 400
                
            baseline_vals['prev_day_1515_close_ha'] = float(last_row[prev_day_col]) if not pd.isna(last_row[prev_day_col]) else 0.0
            baseline_vals['today_0915_close_ha'] = float(last_row[today_0915_col]) if not pd.isna(last_row[today_0915_col]) else 0.0
            
            # 2. Parse candles from 5m_Live_Candles
            # Find time and OHLC columns for option type
            time_col = next((c for c in df_candles.columns if 'Time' in c), df_candles.columns[0])
            open_col = f"{option_type.capitalize()} Open"
            high_col = f"{option_type.capitalize()} High"
            low_col = f"{option_type.capitalize()} Low"
            close_col = f"{option_type.capitalize()} Close"
            
            required_cols = [open_col, high_col, low_col, close_col]
            for col in required_cols:
                if col not in df_candles.columns:
                    return jsonify({"error": f"Could not find column '{col}' in 5m_Live_Candles sheet."}), 400
                    
            for _, row in df_candles.iterrows():
                if pd.isna(row[time_col]) or pd.isna(row[open_col]):
                    continue
                # Time format handling (could be datetime object or string)
                raw_time = row[time_col]
                if hasattr(raw_time, 'strftime'):
                    time_str = raw_time.strftime('%H:%M')
                else:
                    time_str = str(raw_time).strip()
                    if ' ' in time_str:  # strip date if datetime string
                        time_str = time_str.split(' ')[-1]
                    if len(time_str.split(':')) > 2:  # strip seconds
                        time_str = ':'.join(time_str.split(':')[:2])
                        
                candles.append({
                    "time": time_str,
                    "open": float(row[open_col]),
                    "high": float(row[high_col]),
                    "low": float(row[low_col]),
                    "close": float(row[close_col])
                })
        else:
            # Manual mode
            import json
            manual_candles_raw = request.form.get('candles', '[]')
            try:
                candles = json.loads(manual_candles_raw)
            except Exception:
                return jsonify({"error": "Invalid candles JSON format."}), 400
                
            baseline_vals['prev_day_1515_close_ha'] = float(request.form.get('prev_day_1515_close_ha', 0))
            baseline_vals['today_0915_close_ha'] = float(request.form.get('today_0915_close_ha', 0))
            
        if not candles:
            return jsonify({"error": "No 5-minute candle data found or provided."}), 400
            
        # Slicing Rule: Completely ignore the 9:15 AM and 9:20 AM candles
        filtered_candles = []
        for c in candles:
            time_str = str(c['time']).strip()
            if ' ' in time_str:
                time_str = time_str.split(' ')[-1]
            if len(time_str.split(':')) > 2:
                time_str = ':'.join(time_str.split(':')[:2])
            if time_str in ['09:15', '09:20']:
                continue
            c_copy = c.copy()
            c_copy['time'] = time_str
            filtered_candles.append(c_copy)
        candles = filtered_candles
        
        if not candles:
            return jsonify({"error": "No candle data remaining after time filtering."}), 400
            
        # --- Backtest Simulator Logic ---
        
        # 1. Setup Candle Scanner: Monitor candles for a candle completely below baseline (High < Baseline)
        setup_candle_idx = None
        locked_sl = None
        for i, c in enumerate(candles):
            if c['high'] < (manual_baseline if manual_baseline is not None else max(baseline_vals.get('prev_day_1515_close_ha', 0.0), baseline_vals.get('today_0915_close_ha', 0.0))):
                setup_candle_idx = i
                locked_sl = c['low']
                break
        
        # 1. Baseline Calculation
        computed_baseline = max(baseline_vals.get('prev_day_1515_close_ha', 0.0), baseline_vals.get('today_0915_close_ha', 0.0))
        final_baseline = manual_baseline if manual_baseline is not None else computed_baseline
        
        # 2. Entry Detection
        first_breach_idx = None
        entry_idx = None
        entry_price = None
        entry_time = None
        entry_type = "auto"
        
        # Auto Entry detection
        for i, c in enumerate(candles):
            high_val = max(c['open'], c['high'], c['low'], c['close'])
            if high_val > final_baseline:
                if first_breach_idx is None:
                    first_breach_idx = i
                elif entry_idx is None:
                    entry_idx = i
                    entry_price = c['open']
                    entry_time = c['time']
                    break
                    
        # Fallback logic if no 2nd breach
        if first_breach_idx is not None and entry_idx is None:
            c = candles[first_breach_idx]
            entry_idx = first_breach_idx
            entry_price = max(c['open'], c['high'], c['low'], c['close'])
            entry_time = c['time']
            entry_type = "fallback"
            
        # Apply manual entry override
        if manual_entry_time is not None:
            # Find override index
            override_idx = next((i for i, c in enumerate(candles) if c['time'] == manual_entry_time), None)
            if override_idx is not None:
                entry_idx = override_idx
                entry_price = manual_entry_price if manual_entry_price is not None else candles[override_idx]['open']
                entry_time = manual_entry_time
                entry_type = "override"
        elif manual_entry_price is not None and entry_idx is not None:
            entry_price = manual_entry_price
            entry_type = "override"

        # 3. Target Calculation
        target_val = None
        target_hit_idx = None
        target_hit_time = None
        target_exit_price = None
        
        if entry_idx is not None:
            entry_ref_high = candles[entry_idx]['high']
            target_val = manual_target if manual_target is not None else (entry_ref_high + target_points)
            
            # Scan forward from entry_idx (inclusive)
            for i in range(entry_idx, len(candles)):
                c = candles[i]
                if c['high'] >= target_val:
                    target_hit_idx = i
                    target_hit_time = c['time']
                    target_exit_price = target_val
                    break
                    
        # 5. Stop-Loss Calculation
        sl_ref = manual_sl if manual_sl is not None else (locked_sl if locked_sl is not None else final_baseline)
        sl_hit_idx = None
        sl_hit_time = None
        sl_exit_price = None
        
        if entry_idx is not None:
            # If manual stop loss is specified, we check all candles. If automatic locked stop loss is used, it is only active after the setup candle completes.
            if manual_sl is not None or setup_candle_idx is not None:
                start_scan_idx = max(entry_idx, (setup_candle_idx + 1) if manual_sl is None else entry_idx)
                for i in range(start_scan_idx, len(candles)):
                    c = candles[i]
                    trigger_val = c['low'] if sl_type == 'low' else c['close']
                    if trigger_val < sl_ref:
                        sl_hit_idx = i
                        sl_hit_time = c['time']
                        sl_exit_price = min(c['open'], c['high'], c['low'], c['close'])
                        break
                    
        # Apply manual exit override if provided
        forced_exit_idx = None
        if manual_exit_time is not None:
            forced_exit_idx = next((i for i, c in enumerate(candles) if c['time'] == manual_exit_time), None)
            
        # 5. Outcome determination
        exit_reason = "Open/No Exit"
        exit_time = "Ã¢â‚¬â€"
        exit_price = 0.0
        pnl = 0.0
        outcome_color = "status-open"
        
        if forced_exit_idx is not None:
            exit_reason = "Manual Override"
            exit_time = manual_exit_time
            # For P&L check target override price or SL override price or default to that candle's close
            if manual_target is not None:
                exit_price = manual_target
            elif manual_sl is not None:
                exit_price = manual_sl
            else:
                exit_price = candles[forced_exit_idx]['close']
                
            if entry_price is not None:
                pnl = exit_price - entry_price
                outcome_color = "status-win" if pnl >= 0 else "status-loss"
        else:
            # Compare target vs SL hit times
            if target_hit_idx is not None and sl_hit_idx is not None:
                if target_hit_idx < sl_hit_idx:
                    exit_reason = "Target Hit"
                    exit_time = target_hit_time
                    exit_price = target_exit_price
                    pnl = exit_price - entry_price
                    outcome_color = "status-win"
                else:
                    exit_reason = "Stop Loss Hit"
                    exit_time = sl_hit_time
                    exit_price = sl_exit_price
                    pnl = exit_price - entry_price
                    outcome_color = "status-loss"
            elif target_hit_idx is not None:
                exit_reason = "Target Hit"
                exit_time = target_hit_time
                exit_price = target_exit_price
                pnl = exit_price - entry_price
                outcome_color = "status-win"
            elif sl_hit_idx is not None:
                exit_reason = "Stop Loss Hit"
                exit_time = sl_hit_time
                exit_price = sl_exit_price
                pnl = exit_price - entry_price
                outcome_color = "status-loss"
                
        # Tag candles for coloring
        for i, c in enumerate(candles):
            c['tag'] = ""
            if entry_idx is not None and i == entry_idx:
                c['tag'] = "entry"
            elif forced_exit_idx is not None and i == forced_exit_idx:
                c['tag'] = "exit"
            elif exit_reason == "Target Hit" and target_hit_idx is not None and i == target_hit_idx:
                c['tag'] = "target"
            elif exit_reason == "Stop Loss Hit" and sl_hit_idx is not None and i == sl_hit_idx:
                c['tag'] = "stop_loss"
                
        # Resolve target option symbol from task database
        winner_sym = None
        if task_id:
            with tasks_lock:
                task_data = tasks.get(task_id)
                if task_data:
                    symbols = task_data.get('symbols', [])
                    latest_sym = symbols[-1] if symbols else None
                    if latest_sym:
                        winner_sym = latest_sym.get('call_sym') if option_type == 'call' else latest_sym.get('put_sym')

        # Check settings for auto-pilot
        import os, json
        settings_path = os.path.join(os.path.dirname(__file__), 'settings.json')
        auto_pilot = False
        lot_size = 50
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r', encoding='utf-8') as sf:
                    sdata = json.load(sf)
                    lot_size = int(sdata.get('lot_size', 50))
                    if sdata.get('auto_pilot_enabled') and sdata.get('safety_lock_key') == "I UNDERSTAND THE RISK":
                        auto_pilot = True
            except:
                pass

        # Return response
        signal_events = _build_signal_events(
            candles, final_baseline,
            entry_idx, entry_time, entry_price, entry_type,
            None, exit_reason, exit_time, exit_price,
            target_val, sl_ref, option_type,
            lot_size=lot_size,
            tradingsymbol=winner_sym
        )

        # Trigger auto-pilot execution if enabled
        if auto_pilot and entry_idx is not None:
            _run_autopilot_execution(signal_events, winner_sym, pnl)

            # Update event status to show auto-executed
            for ev in signal_events:
                if ev.get('actionable') and ev.get('time') != 'Ã¢â‚¬â€':
                    ev['status'] = 'Ã¢Å¡Â¡ AUTO-EXECUTED'

        return jsonify({
            "status": "success",
            "baseline_values": baseline_vals,
            "computed_baseline": computed_baseline,
            "final_baseline": final_baseline,
            "entry_time": entry_time,
            "entry_price": entry_price,
            "entry_type": entry_type,
            "target_value": target_val,
            "target_hit_time": target_hit_time,
            "sl_value": sl_ref,
            "sl_hit_time": sl_hit_time,
            "sl_exit_price": sl_exit_price,
            "final_exit_reason": exit_reason,
            "final_exit_time": exit_time,
            "final_exit_price": exit_price,
            "pnl": round(pnl, 2) if entry_price is not None else 0.0,
            "outcome_color": outcome_color,
            "signal_events": signal_events,
            "candles": candles
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _run_autopilot_execution(events, tradingsymbol, pnl):
    """Automatically log events to trade journal if auto-pilot is enabled."""
    import os, json
    from datetime import datetime
    
    path = os.path.join(os.path.dirname(__file__), 'journal.json')
    existing = {'trades': []}
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
        except:
            pass

    # Filter actionable events
    actionable_events = [e for e in events if e.get('actionable') and e.get('time') != 'Ã¢â‚¬â€']
    if not actionable_events:
        return

    # Check if we already logged this session's trade (prevent duplicate entries)
    entry_event = next((e for e in actionable_events if e['event_type'] == 'BUY_SIGNAL'), None)
    if not entry_event:
        return

    already_logged = False
    for t in existing['trades']:
        if t.get('symbol') == (tradingsymbol or "NIFTY-OPTION") and t.get('entry_time') == entry_event['time']:
            already_logged = True
            break

    if not already_logged:
        exit_event = next((e for e in actionable_events if e['event_type'] in ('TARGET_EXIT', 'SL_EXIT') and e.get('status') == 'HIT'), None)
        
        trade_data = {
            "timestamp": datetime.now().isoformat(),
            "symbol": tradingsymbol or "NIFTY-OPTION",
            "action": "AUTO-BUY",
            "price": entry_event['price'],
            "qty": entry_event.get('qty', 50),
            "product": entry_event.get('product', 'MIS'),
            "order_type": entry_event.get('order_type', 'MARKET'),
            "status": "COMPLETED (AUTO-PILOT)",
            "pnl": round(pnl, 2),
            "entry_time": entry_event['time']
        }
        
        if exit_event:
            trade_data["exit_time"] = exit_event['time']
            trade_data["exit_price"] = exit_event['price']
            trade_data["exit_reason"] = exit_event['event_label']
            
        existing['trades'].append(trade_data)
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(existing, f, indent=2)
        except:
            pass

def _build_signal_events(candles, baseline, entry_idx, entry_time, entry_price, entry_type,
                         first_breach_idx, exit_reason, exit_time, exit_price,
                         target_val, sl_val, option_type, lot_size=50, tradingsymbol=None):
    """Build structured signal events for the Order Ticket panel."""
    events = []
    side = option_type.upper() if option_type else 'CE'
    sym_hint = tradingsymbol or f'NIFTY{side[:2]}'

    # 1. Breach Alert (informational only)
    if first_breach_idx is not None and first_breach_idx != entry_idx and first_breach_idx < len(candles):
        c = candles[first_breach_idx]
        events.append({
            'event_type': 'BREACH_ALERT',
            'event_label': '1st Breach',
            'time': c['time'],
            'price': round(max(c['open'], c['high'], c['low'], c['close']), 2),
            'action': None,
            'actionable': False,
            'color': 'warning',
            'status': 'INFO',
            'description': f'Price first crossed baseline ({baseline:.2f})',
            'zerodha_payload': None,
        })

    # 2. BUY Signal (entry)
    if entry_idx is not None and entry_price is not None:
        events.append({
            'event_type': 'BUY_SIGNAL',
            'event_label': 'Entry Signal',
            'time': entry_time,
            'price': entry_price,
            'action': 'BUY',
            'actionable': True,
            'color': 'success',
            'status': 'EXECUTED' if exit_reason != 'Open/No Exit' else 'PENDING',
            'description': f'Entry confirmed ({entry_type}) Ã¢â‚¬â€ baseline {baseline:.2f} breached',
            'order_type': 'MARKET',
            'product': 'MIS',
            'qty': lot_size,
            'zerodha_payload': {
                'tradingsymbol': sym_hint,
                'exchange': 'NFO',
                'transaction_type': 'BUY',
                'quantity': lot_size,
                'order_type': 'MARKET',
                'product': 'MIS',
                'validity': 'DAY',
                'price': 0,
                'trigger_price': 0,
                'tag': 'NiftyOps-Entry',
            },
        })

        # 3. TARGET Exit
        if target_val:
            events.append({
                'event_type': 'TARGET_EXIT',
                'event_label': 'Target Exit',
                'time': exit_time if exit_reason == 'Target Hit' else 'Ã¢â‚¬â€',
                'price': round(target_val, 2),
                'action': 'SELL',
                'actionable': True,
                'color': 'accent',
                'status': 'HIT' if exit_reason == 'Target Hit' else 'PENDING',
                'description': f'Sell at target {target_val:.2f} (+{target_val - entry_price:.2f} pts)',
                'order_type': 'LIMIT',
                'product': 'MIS',
                'qty': lot_size,
                'zerodha_payload': {
                    'tradingsymbol': sym_hint,
                    'exchange': 'NFO',
                    'transaction_type': 'SELL',
                    'quantity': lot_size,
                    'order_type': 'LIMIT',
                    'product': 'MIS',
                    'validity': 'DAY',
                    'price': round(target_val, 2),
                    'trigger_price': 0,
                    'tag': 'NiftyOps-Target',
                },
            })

        # 4. SL Exit
        events.append({
            'event_type': 'SL_EXIT',
            'event_label': 'Stop-Loss Exit',
            'time': exit_time if exit_reason == 'Stop Loss Hit' else 'Ã¢â‚¬â€ ',
            'price': round(sl_val, 2),
            'action': 'SELL',
            'actionable': True,
            'color': 'error',
            'status': 'HIT' if exit_reason == 'Stop Loss Hit' else 'PENDING',
            'description': f'Emergency sell if price drops below {sl_val:.2f}',
            'order_type': 'SL-M',
            'product': 'MIS',
            'qty': lot_size,
            'zerodha_payload': {
                'tradingsymbol': sym_hint,
                'exchange': 'NFO',
                'transaction_type': 'SELL',
                'quantity': lot_size,
                'order_type': 'SL-M',
                'product': 'MIS',
                'validity': 'DAY',
                'price': 0,
                'trigger_price': round(sl_val, 2),
                'tag': 'NiftyOps-SL',
            },
        })

    return events


@app.route('/journal')
def journal_page():
    return render_template('journal.html')


@app.route('/api/journal', methods=['GET'])
def get_journal():
    import os, json
    path = os.path.join(os.path.dirname(__file__), 'journal.json')
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    return jsonify({'trades': []})


@app.route('/api/journal', methods=['POST'])
def add_journal():
    import os, json
    path = os.path.join(os.path.dirname(__file__), 'journal.json')
    data = request.get_json(force=True)
    existing = {'trades': []}
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            existing = json.load(f)
    existing['trades'].append(data)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(existing, f, indent=2)
    return jsonify({'ok': True, 'count': len(existing['trades'])})


@app.route('/api/journal', methods=['DELETE'])
def clear_journal():
    import os, json
    path = os.path.join(os.path.dirname(__file__), 'journal.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump({'trades': []}, f)
    return jsonify({'ok': True})


@app.route('/settings')
def settings_page():
    return render_template('settings.html')


@app.route('/api/settings', methods=['GET'])
def get_settings():
    import os, json
    path = os.path.join(os.path.dirname(__file__), 'settings.json')
    defaults = {
        'lot_size': 50, 'target_points': 25, 'sl_type': 'close',
        'exchange': 'NFO', 'product': 'MIS', 'sound_alerts': False,
        'zerodha_api_key': '', 'zerodha_api_secret': '',
        'auto_pilot_enabled': False, 'safety_lock_key': '', 'max_loss_points': 50
    }
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            saved = json.load(f)
        defaults.update(saved)
    return jsonify(defaults)


@app.route('/api/settings', methods=['POST'])
def save_settings():
    import os, json
    path = os.path.join(os.path.dirname(__file__), 'settings.json')
    data = request.get_json(force=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    return jsonify({'ok': True})


@app.route('/api/set-session', methods=['POST', 'OPTIONS'])
def set_session():
    """Receive a TradingView session token from the bookmarklet or any client."""
    if request.method == 'OPTIONS':
        resp = app.make_default_options_response()
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp
    data = request.get_json(force=True) or {}
    # Accept both key names for compatibility
    session_token = (data.get('session_token') or data.get('session_id') or '').strip()

    # Ã¢â€â‚¬Ã¢â€â‚¬ VALIDATION Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    # Check if the user accidentally pasted the bookmarklet code itself
    if session_token and (session_token.startswith('javascript:') or 'function(' in session_token or 'document.cookie' in session_token or len(session_token) > 120):
        error_msg = (
            "Invalid session token. You pasted the bookmarklet code itself! "
            "Please drag the bookmarklet button to your bookmarks bar, open TradingView.com in "
            "a new tab (where you are logged in), and then click the bookmarklet from your bookmarks bar "
            "to automatically capture the session token."
        )
        resp = jsonify({'ok': False, 'error': error_msg})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp, 400

    import os as _os
    env_path = _os.path.join(_os.path.dirname(__file__), '.env')
    lines = []
    found = False
    if _os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        for i, line in enumerate(lines):
            if line.startswith('TV_SESSION_ID='):
                if session_token:
                    lines[i] = f'TV_SESSION_ID={session_token}\n'
                else:
                    lines.pop(i)  # Remove the line to clear session
                found = True
                break
    if not found and session_token:
        lines.append(f'TV_SESSION_ID={session_token}\n')
    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    # Remove cached JWT so a fresh one is auto-fetched for the new session
    lines = [l for l in lines if not l.startswith('TV_JWT_TOKEN=')]
    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    _flush_tv_pool()  # Force fresh TV connection with new session
    msg = 'Session saved! Full options access enabled.' if session_token else 'Session cleared.'
    resp = jsonify({'ok': True, 'message': msg})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp


# Global cache for session validation to avoid synchronous blocking requests
_session_cache_lock = threading.Lock()
_session_cache = {
    'token': None,
    'valid': False,
    'timestamp': 0,
    'checking': False
}

def _bg_validate_session(cookie_val):
    global _session_cache
    import requests as _req, time as _time
    try:
        r = _req.get(
            'https://www.tradingview.com/chart/',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124',
                'Cookie': f'sessionid={cookie_val}'
            },
            timeout=4,
            allow_redirects=True
        )
        valid = 'is-authenticated' in r.text and 'is-not-authenticated' not in r.text
    except Exception:
        valid = True  # Network issue or timeout, default to True so we don't block
    with _session_cache_lock:
        _session_cache['valid'] = valid
        _session_cache['timestamp'] = _time.time()
        _session_cache['checking'] = False



@app.route('/api/set-jwt', methods=['POST', 'OPTIONS'])
def set_jwt_token():
    """Save a TradingView JWT auth token to .env for NFO options data access.

    How to get your JWT token:
    1. Open TradingView.com in Chrome (logged in)
    2. Press F12 > Network tab > filter by "tradingview.com"
    3. Reload the page, find any authenticated XHR request
    4. Look for the 'Authorization: Bearer eyJ...' header value
       OR open Application > Local Storage > tradingview.com > find auth_token key
    5. Copy the JWT (starts with eyJ) and paste it here.
    """
    if request.method == 'OPTIONS':
        resp = app.make_default_options_response()
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp

    data = request.get_json(force=True) or {}
    jwt_token = (data.get('jwt_token') or data.get('token') or '').strip()

    if not jwt_token:
        return jsonify({'ok': False, 'error': 'No JWT token provided'}), 400

    # Basic JWT validation: must start with eyJ (base64 encoded JSON header)
    if not jwt_token.startswith('eyJ'):
        return jsonify({'ok': False, 'error': 'Invalid JWT format. Token must start with eyJ...'}), 400

    import os as _os
    env_path = _os.path.join(_os.path.dirname(__file__), '.env')
    lines = []
    found = False
    if _os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        for i, line in enumerate(lines):
            if line.startswith('TV_JWT_TOKEN='):
                lines[i] = f'TV_JWT_TOKEN={jwt_token}\n'
                found = True
                break
    if not found:
        lines.append(f'TV_JWT_TOKEN={jwt_token}\n')
    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    _flush_tv_pool()  # Force fresh TV connection with new JWT
    resp = jsonify({'ok': True, 'message': f'JWT token saved (length={len(jwt_token)}). NFO options access enabled!'})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.route('/api/session-status', methods=['GET'])
def session_status():
    """Return whether a TV session ID is saved and still valid against TradingView."""
    import os as _os, time as _time
    env_path = _os.path.join(_os.path.dirname(__file__), '.env')
    session_val = ''
    if _os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('TV_SESSION_ID='):
                    session_val = line.split('=', 1)[1].strip()
                    break

    if not session_val:
        return jsonify({'has_session': False, 'session_preview': '', 'session_valid': False,
                        'status_msg': 'No session saved'})

    # Strip 'sessionid=' prefix if present
    cookie_val = session_val
    if cookie_val.startswith('sessionid='):
        cookie_val = cookie_val[len('sessionid='):]

    now = _time.time()
    trigger_check = False
    with _session_cache_lock:
        if (_session_cache['token'] != cookie_val) or (now - _session_cache['timestamp'] > 180):
            if not _session_cache['checking']:
                _session_cache['checking'] = True
                _session_cache['token'] = cookie_val
                trigger_check = True

    if trigger_check:
        t = threading.Thread(target=_bg_validate_session, args=(cookie_val,), daemon=True)
        t.start()
        # Brief pause to allow the background check a chance to run immediately
        _time.sleep(0.05)

    with _session_cache_lock:
        session_valid = _session_cache['valid']

    preview = (cookie_val[:8] + '***') if session_val else ''
    status_msg = 'Session active' if session_valid else 'Session expired — please refresh'
    return jsonify({
        'has_session': bool(session_val),
        'session_preview': preview,
        'session_valid': session_valid,
        'status_msg': status_msg
    })


@app.route('/api/flush-pool', methods=['POST'])
def flush_pool():
    """Clear the TV connection pool so next extraction starts fresh."""
    _flush_tv_pool()
    return jsonify({'ok': True, 'message': 'Connection pool flushed. Next extraction will re-authenticate.'})



@app.route('/api/auto-extract', methods=['POST'])
def auto_extract():
    """Attempt direct SQLite database cookie extraction for session ID."""
    import os, json, base64, shutil, tempfile, sqlite3
    
    profiles = [
        {
            'name': 'Comet',
            'cookies': os.path.expanduser(r'~\AppData\Local\Perplexity\Comet\User Data\Default\Network\Cookies'),
            'state': os.path.expanduser(r'~\AppData\Local\Perplexity\Comet\User Data\Local State')
        },
        {
            'name': 'Chrome',
            'cookies': os.path.expanduser(r'~\AppData\Local\Google\Chrome\User Data\Default\Network\Cookies'),
            'state': os.path.expanduser(r'~\AppData\Local\Google\Chrome\User Data\Local State')
        },
        {
            'name': 'Edge',
            'cookies': os.path.expanduser(r'~\AppData\Local\Microsoft\Edge\User Data\Default\Network\Cookies'),
            'state': os.path.expanduser(r'~\AppData\Local\Microsoft\Edge\User Data\Local State')
        }
    ]
    
    def decrypt_val(enc_val, key):
        try:
            if enc_val[:3] in (b'v10', b'v11'):
                from Crypto.Cipher import AES
                iv = enc_val[3:15]
                payload = enc_val[15:]
                cipher = AES.new(key, AES.MODE_GCM, nonce=iv)
                return cipher.decrypt(payload[:-16]).decode('utf-8', errors='replace')
            import win32crypt
            return win32crypt.CryptUnprotectData(enc_val, None, None, None, 0)[1].decode('utf-8', errors='replace')
        except:
            return None

    def get_key(state_path):
        try:
            with open(state_path, 'r', encoding='utf-8') as f:
                state = json.load(f)
            enc_key = base64.b64decode(state['os_crypt']['encrypted_key'])[5:]
            import win32crypt
            return win32crypt.CryptUnprotectData(enc_key, None, None, None, 0)[1]
        except:
            return None

    found_session = None
    locked_browsers = []
    
    for p in profiles:
        c_path = p['cookies']
        s_path = p['state']
        if not os.path.exists(c_path):
            alt = c_path.replace(r'\Network\Cookies', r'\Cookies')
            if os.path.exists(alt):
                c_path = alt
                
        if os.path.exists(c_path):
            tmp = os.path.join(tempfile.gettempdir(), f"_ax_cookies_{p['name']}.db")
            try:
                # Bypass file lock
                shutil.copy2(c_path, tmp)
            except PermissionError:
                locked_browsers.append(p['name'])
                continue
            except Exception:
                continue
                
            try:
                key = get_key(s_path)
                conn = sqlite3.connect(tmp)
                cursor = conn.cursor()
                cursor.execute("SELECT name, value, encrypted_value FROM cookies WHERE host_key LIKE '%tradingview.com%'")
                rows = cursor.fetchall()
                conn.close()
                for name, val, enc_val in rows:
                    if name == 'sessionid':
                        if val:
                            found_session = val
                        elif enc_val and key:
                            found_session = decrypt_val(enc_val, key)
                        break
            except Exception:
                pass
            finally:
                if os.path.exists(tmp):
                    try: os.remove(tmp)
                    except: pass
            if found_session:
                break
                
    if found_session:
        env_path = os.path.join(os.path.dirname(__file__), '.env')
        lines = []
        found_in_env = False
        if os.path.exists(env_path):
            with open(env_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            for i, line in enumerate(lines):
                if line.startswith('TV_SESSION_ID='):
                    lines[i] = f'TV_SESSION_ID={found_session}\n'
                    found_in_env = True
                    break
        if not found_in_env:
            lines.append(f'TV_SESSION_ID={found_session}\n')
        with open(env_path, 'w', encoding='utf-8') as f:
            f.writelines(lines)
        return jsonify({'ok': True, 'message': 'Session token captured successfully!'})
        
    if locked_browsers:
        msg = (
            "Your browser cookies database is currently locked because Chrome/Edge/Comet is running.\n\n"
            "To capture the session ID automatically WITHOUT closing any browsers:\n"
            "👉 Please click the purple '⚡ Auto-Login (Bypass)' button on the main screen!\n\n"
            "This will open a temporary browser window, sign you in, and capture the session automatically without any locks."
        )
        return jsonify({
            'ok': False, 
            'error': 'LOCKED', 
            'message': msg
        })
        
    return jsonify({
        'ok': False, 
        'error': 'NOT_FOUND', 
        'message': 'TradingView login cookie not found. Please log in to tradingview.com in Chrome, Edge, or Comet first!'
    })


@app.route('/session-helper')
def session_helper():
    bookmarklet_js = """javascript:(function(){
  var sid = document.cookie.split(';').map(c=>c.trim()).filter(c=>c.startsWith('sessionid=')).map(c=>c.split('=')[1])[0];
  if(!sid){alert('TradingView sessionid not found. Make sure you are logged in on tradingview.com first!');return;}
  fetch('http://127.0.0.1:5005/api/set-session',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({session_id:sid})})
  .then(r=>r.json()).then(d=>{alert(d.ok?'Session captured! Go back to http://127.0.0.1:5005 and click Extract.':'Error: '+d.error);})
  .catch(e=>{alert('Error: '+e);});
})();"""
    return render_template('session_helper.html', bookmarklet_js=bookmarklet_js)


# â”€â”€â”€ Headless/Selenium Automated Capture Endpoints â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
headless_login_state = {
    'status': 'idle',
    'message': ''
}
driver_to_close = None


def _save_session_to_env(sid):
    """Persist a captured TV sessionid to .env permanently and flush the connection pool."""
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    lines = []
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    found = False
    for i, line in enumerate(lines):
        if line.strip().startswith('TV_SESSION_ID='):
            lines[i] = f'TV_SESSION_ID={sid}\n'
            found = True
            break
    if not found:
        lines.append(f'TV_SESSION_ID={sid}\n')
    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    # Flush pooled connections so next extract uses the new session
    _flush_tv_pool()



def _build_chrome_options(use_profile=False, profile_path=None):
    """Build ChromeOptions with anti-detection and standard stability settings."""
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    opts = ChromeOptions()
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1100,900")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-default-apps")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option('useAutomationExtension', False)
    if use_profile and profile_path and os.path.exists(profile_path):
        opts.add_argument(f'--user-data-dir={profile_path}')
    return opts


def run_selenium_login_thread(tv_user=None, tv_pass=None):
    """
    Auto-login strategy (in order):
    1. Connect to existing Chrome via remote debugging port (if active).
    2. Fresh Chrome (uses temporary isolated profile in project folder to prevent hanging).
    """
    global driver_to_close
    headless_login_state['status'] = 'running'
    headless_login_state['message'] = 'Launching Chrome browser...'

    driver = None
    used_profile = False

    # Strategy 0: Connect to existing Chrome via remote debugging port (only if port 9222 is listening)
    is_9222_open = False
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.5)
        s.connect(('127.0.0.1', 9222))
        s.close()
        is_9222_open = True
    except Exception:
        is_9222_open = False

    if is_9222_open:
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options as ChromeOptions
            opts0 = ChromeOptions()
            opts0.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
            driver = webdriver.Chrome(options=opts0)
            used_profile = True
            headless_login_state['message'] = 'Connected to existing Chrome session!'
        except Exception:
            driver = None
            used_profile = False

    # Strategy 1: Launch fresh Chrome with a temporary isolated user-data-dir in project folder
    if driver is None:
        try:
            import os, shutil
            from selenium import webdriver
            project_dir = os.path.dirname(os.path.abspath(__file__))
            tmp_user_data = os.path.join(project_dir, 'temp_chrome_profile')
            
            # Clean directory if it exists to ensure fresh start
            if os.path.exists(tmp_user_data):
                try:
                    shutil.rmtree(tmp_user_data, ignore_errors=True)
                except Exception:
                    pass
            os.makedirs(tmp_user_data, exist_ok=True)
            
            opts = _build_chrome_options(use_profile=True, profile_path=tmp_user_data)
            driver = webdriver.Chrome(options=opts)
            headless_login_state['message'] = 'Chrome launched (fresh session).'
        except Exception as e_chrome:
            try:
                from selenium import webdriver
                from selenium.webdriver.edge.options import Options as EdgeOptions
                opts_e = EdgeOptions()
                opts_e.add_argument("--disable-gpu")
                opts_e.add_argument("--no-sandbox")
                opts_e.add_argument("--window-size=1100,900")
                opts_e.add_argument("--disable-blink-features=AutomationControlled")
                opts_e.add_experimental_option("excludeSwitches", ["enable-automation"])
                opts_e.add_experimental_option('useAutomationExtension', False)
                driver = webdriver.Edge(options=opts_e)
                headless_login_state['message'] = 'Edge launched (Chrome fallback).'
            except Exception as e2:
                headless_login_state['status'] = 'failed'
                headless_login_state['message'] = f'Could not launch Chrome or Edge: {e2}'
                return

    driver_to_close = driver
    found = False

    try:
        import time as _t
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium.webdriver.common.keys import Keys

        # Mask navigator.webdriver
        try:
            driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                'source': (
                    "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
                    "window.chrome = {runtime: {}};"
                )
            })
        except Exception:
            pass

        # â”€â”€ Profile mode: navigate to TV and check for existing session â”€â”€â”€â”€â”€â”€â”€â”€
        if used_profile:
            headless_login_state['message'] = 'Opening TradingView to check existing session...'
            driver.get('https://www.tradingview.com/')
            _t.sleep(4)

            # Check cookies immediately
            for ck in driver.get_cookies():
                if ck.get('name') == 'sessionid' and 'tradingview.com' in ck.get('domain', ''):
                    _save_session_to_env(ck['value'])
                    headless_login_state['status'] = 'success'
                    headless_login_state['message'] = 'Session captured from your Chrome profile! Ready to extract.'
                    found = True
                    return  # finally closes driver

            # Not found in profile â€” navigate to sign-in
            headless_login_state['message'] = 'Not logged in via profile. Opening sign-in page...'
            driver.get('https://www.tradingview.com/accounts/signin/')
            _t.sleep(3)
        else:
            headless_login_state['message'] = 'Opening TradingView sign-in page...'
            driver.get('https://www.tradingview.com/accounts/signin/')
            _t.sleep(3)

        # Check fast-path: already logged in on sign-in page redirect
        for ck in driver.get_cookies():
            if ck.get('name') == 'sessionid' and 'tradingview.com' in ck.get('domain', ''):
                _save_session_to_env(ck['value'])
                headless_login_state['status'] = 'success'
                headless_login_state['message'] = 'Already signed in - session captured!'
                found = True
                return

        # â”€â”€ Auto-fill credentials â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Confirmed working (2026-07-16 live probe):
        # - Email button: ActionChains click (direct .click() crashes ChromeDriver)
        # - Fields: name=id_username, name=id_password
        # - Submit: Keys.RETURN on password field
        headless_login_state['message'] = 'Waiting for TradingView sign-in page...'
        try:
            email_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Email']"))
            )
            headless_login_state['message'] = 'Clicking Email login option...'
            _t.sleep(0.8)
            ActionChains(driver).move_to_element(email_btn).pause(0.4).click().perform()
            _t.sleep(2)
        except Exception as e_btn:
            headless_login_state['message'] = f'Email btn note ({type(e_btn).__name__}) - trying form...'

        # Find fields quickly in parallel (max 8s wait, avoids 32s sequential timeout)
        u_input = None
        p_input = None
        start_find = _t.time()
        while _t.time() - start_find < 8:
            for sel in ["input[name='id_username']", "input[name='username']", "input[name='email']", "#id_username"]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el and el.is_displayed():
                        u_input = el
                        break
                except:
                    pass
            for sel in ["input[name='id_password']", "input[type='password']", "input[name='password']", "#id_password"]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el and el.is_displayed():
                        p_input = el
                        break
                except:
                    pass
            if u_input or p_input:
                break
            _t.sleep(0.5)

        if (u_input or p_input) and (tv_user or tv_pass):
            if u_input and tv_user:
                headless_login_state['message'] = 'Entering username...'
                u_input.clear()
                u_input.send_keys(tv_user)
                _t.sleep(0.5)
            
            # Re-detect password input if not found in first pass (in case username entry triggered it)
            if not p_input:
                for sel in ["input[name='id_password']", "input[type='password']", "input[name='password']", "#id_password"]:
                    try:
                        el = driver.find_element(By.CSS_SELECTOR, sel)
                        if el and el.is_displayed():
                            p_input = el
                            break
                    except:
                        pass

            if p_input and tv_pass:
                headless_login_state['message'] = 'Entering password...'
                p_input.clear()
                p_input.send_keys(tv_pass)
                _t.sleep(0.5)
                headless_login_state['message'] = 'Submitting credentials... (complete any verification in the browser window)'
                p_input.send_keys(Keys.RETURN)
            else:
                headless_login_state['message'] = 'Password field not found. Please log in manually.'
        elif not tv_user or not tv_pass:
            headless_login_state['message'] = 'Credentials not configured. Please sign in manually in the browser window.'
        else:
            headless_login_state['message'] = 'Login fields not detected. Please sign in manually in the browser window.'

        # â”€â”€ Poll for sessionid up to 10 minutes (user may need to do OTP/CAPTCHA) â”€â”€
        MAX_WAIT = 600  # 10 minutes
        start_time = _t.time()
        while _t.time() - start_time < MAX_WAIT:
            try:
                _ = driver.window_handles
            except Exception:
                headless_login_state['status'] = 'failed'
                headless_login_state['message'] = 'Browser was closed before session was captured.'
                break

            sid = None
            for ck in driver.get_cookies():
                if ck.get('name') == 'sessionid' and 'tradingview.com' in ck.get('domain', ''):
                    sid = ck['value']
                    break

            if sid:
                _save_session_to_env(sid)

                # ── Also extract JWT from localStorage for NFO options access ──
                jwt_token = ""
                try:
                    # TradingView stores auth info in localStorage
                    jwt_scripts = [
                        # Try localStorage for auth_token
                        "return window.localStorage.getItem('auth_token');",
                        # Try user data object
                        "try { var d=JSON.parse(window.localStorage.getItem('tv_user_logged_in_data')||'{}'); return d.auth_token||d.token||''; } catch(e){ return ''; }",
                        # Try from cookies via JS
                        "try { var m=document.cookie.match(/auth_token=([^;]+)/); return m?m[1]:''; } catch(e){ return ''; }",
                    ]
                    for script in jwt_scripts:
                        try:
                            val = driver.execute_script(script)
                            if val and str(val).startswith('eyJ') and len(str(val)) > 50:
                                jwt_token = str(val).strip()
                                break
                        except Exception:
                            pass

                    # Fallback: scrape JWT from page source
                    if not jwt_token:
                        import re as _re
                        page_src = driver.page_source
                        m = _re.search(r'"auth_token"\s*:\s*"(eyJ[A-Za-z0-9_\-\.]+)"', page_src)
                        if m:
                            jwt_token = m.group(1)
                except Exception as e_jwt:
                    pass

                if jwt_token:
                    # Save JWT to .env
                    import os as _os2
                    _env2 = _os2.path.join(_os2.path.dirname(__file__), '.env')
                    _lines2 = []
                    _found_jwt = False
                    if _os2.path.exists(_env2):
                        with open(_env2, 'r', encoding='utf-8') as _f2:
                            _lines2 = _f2.readlines()
                        for _i2, _l2 in enumerate(_lines2):
                            if _l2.startswith('TV_JWT_TOKEN='):
                                _lines2[_i2] = f'TV_JWT_TOKEN={jwt_token}\n'
                                _found_jwt = True
                                break
                    if not _found_jwt:
                        _lines2.append(f'TV_JWT_TOKEN={jwt_token}\n')
                    with open(_env2, 'w', encoding='utf-8') as _f2:
                        _f2.writelines(_lines2)
                    _flush_tv_pool()
                    headless_login_state['status'] = 'success'
                    headless_login_state['message'] = 'Session + JWT captured! Full NFO options access enabled. Ready to extract.'
                else:
                    _flush_tv_pool()
                    headless_login_state['status'] = 'success'
                    headless_login_state['message'] = 'Session captured! Ready to extract. (JWT not found — index data only; NFO options may be limited)'

                found = True
                break

            elapsed = int(_t.time() - start_time)
            remaining = MAX_WAIT - elapsed
            headless_login_state['message'] = (
                f'Waiting for login... ({elapsed}s elapsed, {remaining}s remaining). '
                'Browser is open - complete any verification in the browser window.'
            )
            _t.sleep(1)

        if not found and headless_login_state['status'] == 'running':
            headless_login_state['status'] = 'failed'
            headless_login_state['message'] = 'Login timed out (10 minutes). Please try again.'

    except Exception as e:
        headless_login_state['status'] = 'failed'
        headless_login_state['message'] = f'Login error: {str(e)}'
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        driver_to_close = None


@app.route('/api/headless-login', methods=['POST'])
def start_headless_login():
    data = request.get_json(force=True) or {}
    tv_user = data.get('username', '').strip()
    tv_pass = data.get('password', '').strip()
    headless_login_state['status'] = 'idle'
    headless_login_state['message'] = ''
    t = threading.Thread(target=run_selenium_login_thread, args=(tv_user, tv_pass), daemon=True)
    t.start()
    return jsonify({'ok': True, 'message': 'Auto-login started'})


@app.route('/api/headless-login-status', methods=['GET'])
def get_headless_login_status():
    return jsonify(headless_login_state)


def open_browser():
    """Open the app in the default browser using robust native shell command."""
    if 'ANTIGRAVITY_AGENT' in os.environ:
        print("[BOOT] Antigravity Sandbox environment detected. Skipping browser auto-opening.")
        return
    time.sleep(1.5)
    url = "http://127.0.0.1:5005"
    print(f"[BOOT] Force opening Chrome to {url}...")
    import subprocess
    try:
        # Use cmd /c start chrome to bypass session isolation on Windows
        subprocess.run(['cmd.exe', '/c', 'start', 'chrome', url], shell=True, check=True)
    except Exception:
        try:
            subprocess.run(['cmd.exe', '/c', 'start', url], shell=True)
        except Exception:
            import webbrowser
            webbrowser.open(url)


def auto_login_on_startup():
    """Automatically check and trigger TV session capture on server startup."""
    if 'ANTIGRAVITY_AGENT' in os.environ:
        print("[BOOT] Antigravity Sandbox environment detected. Skipping startup auto-login to prevent GUI thread hanging.")
        return
    try:
        import time as _t
        _t.sleep(2.0)  # Wait for Flask to initialize
        
        env_file = os.path.join(os.path.dirname(__file__), '.env')
        username = ""
        password = ""
        session_id = ""
        
        if os.path.exists(env_file):
            with open(env_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('TV_USERNAME='):
                        username = line.split('=', 1)[1].strip()
                    elif line.startswith('TV_PASSWORD='):
                        password = line.split('=', 1)[1].strip()
                    elif line.startswith('TV_SESSION_ID='):
                        session_id = line.split('=', 1)[1].strip()
        
        # If credentials exist but session ID is missing, run auto-login immediately
        if username and password and not session_id:
            print(f"[BOOT] Session ID missing. Auto-triggering TradingView login for user '{username}'...")
            t = threading.Thread(target=run_selenium_login_thread, args=(username, password), daemon=True)
            t.start()
        else:
            print("[BOOT] Session ID already exists or credentials missing. Skipping startup auto-login.")
    except Exception as e:
        print(f"[BOOT] Startup auto-login error: {e}")


if __name__ == '__main__':
    # Launch browser on startup in background thread
    threading.Thread(target=open_browser, daemon=True).start()
    # (Disabled startup auto-login thread to prevent lock conflict hangs; trigger manually via UI Settings if needed)
    # threading.Thread(target=auto_login_on_startup, daemon=True).start()
    
    # Run server locally
    print("[BOOT] Starting TradingView Extractor Server on http://127.0.0.1:5005...")
    app.run(host='0.0.0.0', port=5005, debug=False, threaded=True)
